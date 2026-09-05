"""Excel 파일 파싱 모듈 (Import Studio S2 BE)"""
from __future__ import annotations

import re
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def col_letter_to_index(col: str) -> int:
    """'A열' → 0, 'B열' → 1, ...

    openpyxl.utils.column_index_from_string은 1-indexed를 반환하므로 1 차감.
    """
    from openpyxl.utils import column_index_from_string

    letter = col.replace("열", "").strip().upper()
    if not letter:
        raise ValueError(f"열 이름을 파싱할 수 없습니다: {col!r}")
    return column_index_from_string(letter) - 1  # 0-indexed


def file_sha256(file_path: Path) -> str:
    """Return a stable content identity used to reject stale previews."""
    digest = hashlib.sha256()
    with file_path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _detect_header_row(ws, col_indices: dict[str, int]) -> int:
    """Find the mapped header row without assuming it is the first row.

    QA spreadsheets commonly contain a title row above the real header.  When
    no recognizable header exists we conservatively treat row 1 as the header,
    preserving the legacy simple-sheet behaviour.
    """
    aliases = {
        "tc_id": ("scenario id", "test scenario id", "tc_id", "테스트케이스 id"),
        "title": ("title", "summary", "테스트 시나리오", "테스트 항목"),
        "steps": ("steps", "테스트 절차"),
        "expected": ("expected", "기대결과"),
    }
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        matches = 0
        for field, expected in aliases.items():
            idx = col_indices.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                continue
            value = re.sub(r"\s+", " ", str(row[idx]).strip().lower())
            if any(alias in value for alias in expected):
                matches += 1
        if matches >= 2:
            return row_idx
    return 1


def parse_sheet(
    file_path: Path,
    sheet_name: str,
    mappings: dict,
    *,
    header_row: int | None = None,
) -> list[dict]:
    """Excel 시트를 읽어 TC 행 목록으로 변환.

    Args:
        file_path: .xlsx 파일 경로
        sheet_name: 읽을 시트 이름
        mappings: {"tc_id": "A열", "title": "B열", "steps": "C열", ...}

    Returns:
        [
            {
                "tc_id": "PL-01",
                "title": "...",
                "steps": "...",
                "expected": "...",
                "priority": "High",
                "tags": [],
                "group": "partner_login",
                "_row": 2,
            },
            ...
        ]

    Notes:
        - 첫 번째 행은 헤더로 처리해 건너뜀.
        - group은 mappings에 있으면 해당 열, 없으면 빈 문자열.
        - tags는 쉼표 구분 문자열을 list로 변환.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}'가 없습니다. 사용 가능: {wb.sheetnames}")

        ws = wb[sheet_name]

        # 컬럼 인덱스 사전 변환 — 잘못된 열 표기를 조기에 잡음
        col_indices: dict[str, int] = {}
        for field, col_expr in mappings.items():
            try:
                col_indices[field] = col_letter_to_index(col_expr)
            except Exception as exc:
                raise ValueError(f"매핑 필드 '{field}'의 열 '{col_expr}' 변환 실패: {exc}") from exc

        rows: list[dict] = []
        effective_header = header_row or _detect_header_row(ws, col_indices)

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx <= effective_header:
                continue

            def _cell(field: str) -> str:
                idx = col_indices.get(field)
                if idx is None:
                    return ""
                if idx >= len(row):
                    return ""
                val = row[idx]
                return str(val).strip() if val is not None else ""

            mapped_values = {field: _cell(field) for field in col_indices}
            tc_id = mapped_values.get("tc_id", "")
            title = mapped_values.get("title", "")
            steps = mapped_values.get("steps", "")
            expected = mapped_values.get("expected", "")

            # Only a completely empty mapped row is formatting noise. Partial
            # rows must reach validation so no user data disappears silently.
            if not any(mapped_values.values()):
                continue

            # tags 처리: 쉼표 구분 문자열 → list
            tags_raw = _cell("tags")
            tags: list[str] = (
                [t.strip() for t in tags_raw.split(",") if t.strip()]
                if tags_raw
                else []
            )

            group = _cell("group") or re.sub(r'\s+', '_', sheet_name.strip().lower())  # mappings에 없으면 시트명을 그룹으로

            entry: dict[str, Any] = {
                "tc_id": tc_id,
                "title": title,
                "precondition": _cell("precondition"),
                "steps": steps,
                "expected": expected,
                "priority": _cell("priority"),
                "tags": tags,
                "group": group,
                "_row": row_idx,
                "_source_file": file_path.name,
                "_source_sheet": sheet_name,
            }
            rows.append(entry)
    finally:
        wb.close()

    return rows


def get_file_metadata(file_path: Path) -> dict:
    """파일 메타데이터 + 시트 목록 반환.

    Returns:
        {
            "sheets": ["Sheet1", "TC목록"],
            "size": 102400,
            "modified": "2026-09-01T10:00:00Z",
        }
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    stat = file_path.stat()
    modified_dt = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)

    return {
        "sheets": sheet_names,
        "size": stat.st_size,
        "modified": modified_dt.strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
