from __future__ import annotations

import json
import os
import time
import urllib.request
from pathlib import Path

import pytest

from tests.import_studio_test_support import (
    dashboard_server,
    materialize_workbooks,
    request_json,
    tree_digest,
    workbook_id,
)


MAPPING = {
    "tc_id": "A열",
    "title": "B열",
    "steps": "C열",
    "expected": "D열",
    "priority": "E열",
    "tags": "F열",
    "group": "G열",
}


@pytest.fixture
def import_project(tmp_path: Path) -> Path:
    project = tmp_path / "project"
    materialize_workbooks(project / "import")
    (project / "testcases").mkdir(parents=True)
    return project


@pytest.fixture
def import_api(import_project: Path):
    with dashboard_server(import_project) as base_url:
        yield base_url


def _all_sources() -> list[dict]:
    return [
        {"file_id": workbook_id("multi_a.xlsx"), "sheet_name": "Login", "mappings": MAPPING},
        {"file_id": workbook_id("multi_a.xlsx"), "sheet_name": "Admin", "mappings": MAPPING},
        {"file_id": workbook_id("multi_b.xlsx"), "sheet_name": "Payment", "mappings": MAPPING},
        {"file_id": workbook_id("multi_b.xlsx"), "sheet_name": "Refund", "mappings": MAPPING},
    ]


def _preview(import_api: str, sources: list[dict] | None = None) -> dict:
    status, body = request_json(
        import_api,
        "POST",
        "/api/import/preview",
        {"sources": sources or _all_sources()},
    )
    assert status == 200, body
    assert body["ok"] is True
    assert body.get("run_id"), body
    assert body["status"] == "preview_ready"
    return body


def test_preview_combines_multiple_files_and_sheets_into_one_run(import_api: str):
    preview = _preview(import_api)

    assert len(preview["sources"]) == 4
    assert sum(preview["summary"].values()) == 6
    assert preview["summary"]["added"] == 6
    assert len(preview["rows"]) == 6
    assert {row["sheet_name"] for row in preview["rows"]} == {
        "Login", "Admin", "Payment", "Refund"
    }
    assert {row["file_name"] for row in preview["rows"]} == {
        "multi_a.xlsx", "multi_b.xlsx"
    }
    assert all(row["source_row"] >= 2 for row in preview["rows"])


def test_preview_reports_invalid_identifier_instead_of_silently_dropping_row(
    import_api: str, import_project: Path
):
    import openpyxl

    path = import_project / "import" / "invalid.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cases"
    sheet.append(["tc_id", "title", "steps", "expected", "priority", "tags", "group"])
    sheet.append(["BAD/ID", "unsafe id", "one", "two", "high", "", "invalid"])
    workbook.save(path)
    workbook.close()

    preview = _preview(import_api, [{
        "file_id": workbook_id("invalid.xlsx"),
        "sheet_name": "Cases",
        "mappings": MAPPING,
    }])

    assert len(preview["rows"]) == 1
    assert preview["summary"]["error"] == 1
    assert preview["rows"][0]["status"] == "error"
    assert preview["rows"][0]["reason_code"] == "INVALID_TC_ID"
    assert "tc_id" in preview["rows"][0]["reason"]

    with urllib.request.urlopen(
        f"{import_api}/api/import/runs/{preview['run_id']}/skipped.csv", timeout=10
    ) as response:
        csv_text = response.read().decode("utf-8-sig")
        assert response.status == 200
        assert "BAD/ID" in csv_text
        assert "reason_code" in csv_text
        assert "INVALID_TC_ID" in csv_text


def test_commit_and_rollback_restore_exact_original_testcase_tree(
    import_api: str, import_project: Path
):
    original = import_project / "testcases" / "login" / "tc_LOGIN_001_old_title.md"
    original.parent.mkdir(parents=True)
    original.write_text(
        "---\nid: LOGIN_001\ndata_key: null\npriority: low\ntags: [legacy]\ntype: structured\n---\n"
        "# old title\n\n## Steps\nold step\n\n## Expected\nold result\n",
        encoding="utf-8",
    )
    before = tree_digest(import_project / "testcases")

    preview = _preview(import_api)
    assert preview["summary"]["updated"] == 1

    status, commit = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 200, commit
    assert commit["ok"] is True
    assert commit["run_id"] == preview["run_id"]
    assert commit["committed"] == 6
    assert tree_digest(import_project / "testcases") != before
    assert not original.exists(), "title-derived filename must be replaced, not duplicated"
    assert list((import_project / "testcases" / "login").glob("tc_LOGIN_001_valid_login.md"))

    status, rollback = request_json(
        import_api,
        "POST",
        f"/api/import/runs/{preview['run_id']}/rollback",
        {},
    )
    assert status == 200, rollback
    assert rollback["ok"] is True
    assert rollback["status"] == "rolled_back"
    assert rollback["verified"] is True
    assert tree_digest(import_project / "testcases") == before
    assert original.exists()


def test_second_target_write_failure_automatically_restores_tree(
    import_api: str, import_project: Path, monkeypatch: pytest.MonkeyPatch
):
    preview = _preview(import_api)
    before = tree_digest(import_project / "testcases")
    original_replace = Path.replace
    target_writes = 0

    def fail_second_testcase_write(source: Path, target: Path):
        nonlocal target_writes
        target_path = Path(target)
        if target_path.suffix == ".md" and "testcases" in target_path.parts:
            target_writes += 1
            if target_writes == 2:
                raise OSError("injected second testcase write failure")
        return original_replace(source, target)

    monkeypatch.setattr(Path, "replace", fail_second_testcase_write)
    status, body = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )

    assert status == 500, body
    assert body["ok"] is False
    assert body["code"] == "COMMIT_RECOVERED"
    assert tree_digest(import_project / "testcases") == before


def test_run_result_is_queryable_after_commit(import_api: str):
    preview = _preview(import_api)
    status, commit = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 200, commit

    status, run = request_json(
        import_api, "GET", f"/api/import/runs/{preview['run_id']}"
    )
    assert status == 200, run
    assert run["ok"] is True
    assert run["run_id"] == preview["run_id"]
    assert run["status"] == "committed"
    assert run["snapshot_id"] == commit["snapshot_id"]


def test_duplicate_commit_is_rejected_without_changing_files(
    import_api: str, import_project: Path
):
    preview = _preview(import_api)
    status, first = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 200, first
    committed_tree = tree_digest(import_project / "testcases")

    status, second = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 409, second
    assert second["ok"] is False
    assert second["code"] == "ALREADY_COMMITTED"
    assert tree_digest(import_project / "testcases") == committed_tree


def test_same_idempotency_key_replays_original_commit_result(import_api: str):
    preview = _preview(import_api)
    payload = {"run_id": preview["run_id"], "idempotency_key": "browser-operation-001"}
    status, first = request_json(import_api, "POST", "/api/import/commit", payload)
    assert status == 200, first

    status, replay = request_json(import_api, "POST", "/api/import/commit", payload)
    assert status == 200, replay
    assert replay["ok"] is True
    assert replay["replayed"] is True
    assert replay["snapshot_id"] == first["snapshot_id"]
    assert replay["committed"] == first["committed"]

    status, conflict = request_json(
        import_api,
        "POST",
        "/api/import/commit",
        {"run_id": preview["run_id"], "idempotency_key": "different-operation"},
    )
    assert status == 409, conflict
    assert conflict["code"] == "ALREADY_COMMITTED"


def test_commit_rejects_source_changed_after_preview(
    import_api: str, import_project: Path
):
    import openpyxl

    preview = _preview(import_api)
    workbook_path = import_project / "import" / "multi_a.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    workbook["Login"].append([
        "LOGIN_003", "late row", "late step", "late result", "low", "", "login"
    ])
    workbook.save(workbook_path)
    workbook.close()
    before = tree_digest(import_project / "testcases")

    status, body = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 409, body
    assert body["ok"] is False
    assert body["code"] == "SOURCE_CHANGED"
    assert tree_digest(import_project / "testcases") == before


def test_commit_rejects_target_changed_after_preview(
    import_api: str, import_project: Path
):
    target = import_project / "testcases" / "login" / "tc_LOGIN_001_old.md"
    target.parent.mkdir(parents=True)
    target.write_text(
        "---\nid: LOGIN_001\n---\n# old\n\n## Steps\nold\n\n## Expected\nold\n",
        encoding="utf-8",
    )
    preview = _preview(import_api)
    target.write_text(target.read_text(encoding="utf-8") + "\nexternal edit\n", encoding="utf-8")
    before = tree_digest(import_project / "testcases")

    status, body = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 409, body
    assert body["ok"] is False
    assert body["code"] == "TARGET_CHANGED"
    assert tree_digest(import_project / "testcases") == before


def test_dead_commit_lock_is_reclaimed(import_api: str, import_project: Path):
    preview = _preview(import_api)
    lock = import_project / "state" / "import_sessions" / ".commit.lock"
    lock.write_text(
        json.dumps({"pid": 99_999_999, "created_at": "2000-01-01T00:00:00Z", "run_id": "dead"}),
        encoding="utf-8",
    )

    status, body = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": preview["run_id"]}
    )
    assert status == 200, body
    assert not lock.exists()


def test_corrupt_lock_requires_staleness_before_reclaim(
    import_api: str, import_project: Path
):
    first_preview = _preview(import_api)
    lock = import_project / "state" / "import_sessions" / ".commit.lock"
    lock.write_text("not-json", encoding="utf-8")

    status, blocked = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": first_preview["run_id"]}
    )
    assert status == 409, blocked
    assert blocked["code"] == "COMMIT_LOCKED"
    assert lock.exists()

    old = time.time() - 301
    os.utime(lock, (old, old))
    status, committed = request_json(
        import_api, "POST", "/api/import/commit", {"run_id": first_preview["run_id"]}
    )
    assert status == 200, committed
    assert not lock.exists()


def test_preview_of_200_rows_completes_within_two_seconds(
    import_api: str, import_project: Path
):
    import openpyxl

    path = import_project / "import" / "two_hundred.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cases"
    sheet.append(["tc_id", "title", "steps", "expected", "priority", "tags", "group"])
    for index in range(200):
        sheet.append([
            f"PERF_{index:03d}", f"case {index}", "perform step", "observe result",
            "medium", "performance", "performance",
        ])
    workbook.save(path)
    workbook.close()

    started = time.perf_counter()
    preview = _preview(import_api, [{
        "file_id": workbook_id("two_hundred.xlsx"),
        "sheet_name": "Cases",
        "mappings": MAPPING,
    }])
    elapsed = time.perf_counter() - started

    assert len(preview["rows"]) == 200
    assert preview["summary"]["added"] == 200
    assert elapsed < 2.0, f"200-row preview took {elapsed:.3f}s"
