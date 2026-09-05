"""
QA Agent Dashboard 서버
프로젝트 루트 또는 어디서든 실행 가능:
  python agents/dashboard/serve.py
"""
from __future__ import annotations

import argparse
import json
import os
import queue
import re
import sys
import threading
import time
import webbrowser
from datetime import datetime
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path

PORT = 8766
HERE = Path(__file__).parent                    # agents/dashboard/
PROJECT_ROOT = HERE.parent.parent               # qa-native/

sys.path.insert(0, str(PROJECT_ROOT / "scripts"))
# .venv site-packages를 sys.path에 추가 (시스템 Python으로 실행해도 패키지 사용 가능)
_venv_sp = PROJECT_ROOT / ".venv" / "lib"
if _venv_sp.exists():
    for _sp in _venv_sp.glob("python*/site-packages"):
        if str(_sp) not in sys.path:
            sys.path.insert(0, str(_sp))
from _python import PYTHON_EXE
from _validators import is_valid_url, is_valid_group_name, is_safe_filename
from _pipeline_registry import (
    Step, ParallelStatus, make_initial_pipeline_state, PIPELINE_STEP_DEFS,
    STEP_COMPAT, PARALLEL_STEP_LABELS,  # P58: 단일 소스에서 임포트
    RESETTABLE_PARALLEL_STATUSES,       # M-4(P121): heal_count 리셋 정책 단일 소스
    STEP_DEF_BY_NAME,                   # M-2(P134): step_labels first-wins 일관성
)
# 상태 파일 경로 + 안전한 쓰기 함수는 _paths.py가 단일 소스다 (#25).
# 예전엔 이 파일이 자체 STATE_PATH 등을 재선언하고 _safe_write_json/
# _safe_update_json을 따로 구현했는데, 그 사본은 (a) 락 획득 실패를 무시하고
# 진행했고 (b) FSM 전이 검증을 안 거쳐 대시보드로 상태를 조작하면 CLI 경로의
# 안전장치가 전부 우회됐다. 이름은 기존 호출부와의 diff를 줄이려고 별칭으로 유지.
from _paths import (
    PIPELINE_STATE as STATE_PATH,
    PARALLEL_STATE as PARALLEL_STATE_PATH,
    QUICK_STATE as QUICK_STATE_PATH,
    RUN_HISTORY as RUN_HISTORY_PATH,
    DISCUSS_STATE as DISCUSS_PATH,
    write_state as _safe_write_json,
    update_state as _safe_update_json,
    reset_state,
    GENERATED_DIR,
    REPORTS_DIR,
    SCREENSHOTS_DIR,
    VIDEOS_DIR,
    HEAL_STATS_PATH,
    FLAKY_TESTS_PATH,
    LOGS_DIR,
    IMPORT_DIR,
    IMPORT_SESSIONS_DIR,
    IMPORT_SNAPSHOTS_DIR,
)
DIALOG_PATH = PROJECT_ROOT / "agents" / "dialog.json"
TEAM_NOTES_PATH = PROJECT_ROOT / "agents" / "team_notes.md"
PENDING_IMPL_PATH = PROJECT_ROOT / "pending_impl.json"
LOGS_DIR.mkdir(exist_ok=True)

ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "http://localhost:8766")
# DNS rebinding 방어: 허용할 Host 헤더 값 목록 (P49)
_allowed_hosts_env = os.environ.get("ALLOWED_HOSTS", "")
ALLOWED_HOSTS = (
    {host for host in re.split(r"[\s,]+", _allowed_hosts_env) if host}
    if _allowed_hosts_env
    else {"localhost:8766", "127.0.0.1:8766"}
)

# Remote mode protects process-spawning and reset endpoints by default. Allowlist
# entries are exact paths unless they end in one ``*``, which means literal prefix.
REMOTE_MODE = os.environ.get("REMOTE_MODE", "").lower() in {
    "1", "true", "yes", "on",
}
REMOTE_API_ALLOWLIST = [
    pattern.strip()
    for pattern in os.environ.get("REMOTE_API_ALLOWLIST", "").split(",")
    if pattern.strip()
]

# ── 서버가 띄운 자식 프로세스 추적 (P61) ────────────────────────
# set[int] → dict[int, Popen] 으로 변경해 liveness 확인(poll()) 가능.
# kill 전 poll()로 이미 종료된 프로세스를 확인해 PID 재사용 오살 위험 감소.
_SPAWNED_PROCS: dict = {}   # dict[int, subprocess.Popen]
_SPAWNED_PIDS_LOCK = threading.Lock()

# P77: 스크립트 태그별 실행 중 프로세스 추적 — 중복 spawn 방지.
# 대시보드 Run 버튼 더블클릭 시 동일 파이프라인이 두 번 기동되어 state 파일 충돌 발생.
# {tag: Popen} — tag는 파이프라인 종류를 나타내는 짧은 문자열.
_SCRIPT_RUNNING: dict[str, object] = {}  # dict[str, subprocess.Popen]


def _register_spawned_proc(proc, tag: str = "") -> None:
    """Popen 객체를 등록하고 죽은 프로세스를 정리한다 (P61).

    P77: tag를 지정하면 동종 프로세스를 _SCRIPT_RUNNING에도 기록해
    _is_script_running()으로 중복 실행을 사전 차단할 수 있다.
    """
    with _SPAWNED_PIDS_LOCK:
        dead = [pid for pid, p in _SPAWNED_PROCS.items() if p.poll() is not None]
        for pid in dead:
            del _SPAWNED_PROCS[pid]
        _SPAWNED_PROCS[proc.pid] = proc
        if tag:
            _SCRIPT_RUNNING[tag] = proc


def _is_script_running(tag: str) -> bool:
    """동종(같은 tag) 프로세스가 이미 실행 중인지 확인 (P77: 중복 spawn 방지)."""
    if not tag:
        return False
    with _SPAWNED_PIDS_LOCK:
        proc = _SCRIPT_RUNNING.get(tag)
        return proc is not None and proc.poll() is None  # type: ignore[union-attr]


# 하위 호환 별칭 — 기존 호출부(proc.pid 전달)가 있을 경우를 위해 유지.
def _register_spawned_pid(pid: int):  # type: ignore[override]
    pass  # 직접 PID만 전달하는 구 호출 경로. _register_spawned_proc를 쓸 것.


def _is_spawned_pid(pid: int) -> bool:
    with _SPAWNED_PIDS_LOCK:
        return pid in _SPAWNED_PROCS

# ── SSE 클라이언트 관리 ────────────────────────────────────────
_sse_clients: list[queue.Queue] = []
_sse_lock = threading.Lock()


def _sse_notify():
    """dialog.json 변경 시 모든 SSE 클라이언트에 알림."""
    with _sse_lock:
        dead = []
        for q in _sse_clients:
            try:
                q.put_nowait("update")
            except queue.Full:
                dead.append(q)
        for q in dead:
            _sse_clients.remove(q)


def _watch_files():
    """dialog.json / state/discuss.json mtime을 0.3초마다 감시."""
    watched = [DIALOG_PATH, DISCUSS_PATH, STATE_PATH, PARALLEL_STATE_PATH,
               QUICK_STATE_PATH]
    last_mtimes = {p: 0.0 for p in watched}
    while True:
        for p in watched:
            try:
                mtime = p.stat().st_mtime if p.exists() else 0.0
                if mtime != last_mtimes[p]:
                    last_mtimes[p] = mtime
                    _sse_notify()
            except Exception:
                pass
        time.sleep(0.3)


# 파일 감시 스레드 시작
threading.Thread(target=_watch_files, daemon=True).start()

TEAM_NOTES_HEADER = (
    "# 팀 결정 사항\n\n"
    "> **독자**: 심의 Agent — 팀 토론 결론 누적. 토론 시 중복 결론 방지 목적으로 참조.\n\n"
    "---\n"
)


def load_json(path: Path):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[Dashboard] JSON 파싱 실패: {path} — {e}")
            return None
    return None


# _safe_write_json/_safe_update_json은 원래 이 파일이 자체 구현했다 (#25).
# 지금은 모듈 상단에서 _paths.write_state/update_state를 같은 이름으로
# import해서 쓴다 — 그래야 락 획득 실패가 조용히 무시되지 않고(TimeoutError로
# 승격) pipeline.json/parallel.json/quick.json 쓰기가 FSM 전이 검증을 받는다.
# (state/pipeline.json·parallel.json을 "init"/""로 되돌리는 리셋 핸들러는
# 예외 — FSM 규칙상 모든 상태에서 init으로 못 돌아가므로 검증을 우회하는
# reset_state()를 별도로 쓴다. 아래 _post_*_reset 참조.)


def parse_conclusion_items(conclusion: str) -> list:
    """결론 마크다운을 투표 가능한 개별 항목으로 파싱."""
    items = []

    # 1) ### 소제목 파싱
    for m in re.finditer(r'^###\s+(.+?)\n([\s\S]*?)(?=^###\s|\Z)', conclusion, re.MULTILINE):
        title = m.group(1).strip()
        body  = re.sub(r'\n?---+\s*$', '', m.group(2)).strip()
        items.append({"id": len(items), "title": title, "text": body, "status": "pending"})

    # 2) 번호 목록 파싱 (1. **title**: body)
    if not items:
        for m in re.finditer(r'^\d+\.\s+(.+)$', conclusion, re.MULTILINE):
            text = m.group(1).strip()
            bold = re.match(r'\*\*(.+?)\*\*[:\s]*(.*)', text)
            title = bold.group(1).strip() if bold else text[:70]
            items.append({"id": len(items), "title": title, "text": text, "status": "pending"})

    # 3) 괄호 번호 패턴 파싱: (1) ... (2) ... 또는 인라인 구분
    if not items:
        parts = re.split(r'\s*\((\d+)\)\s*', conclusion)
        # parts: ['앞부분', '1', '내용1', '2', '내용2', ...]
        if len(parts) >= 3:
            for i in range(1, len(parts) - 1, 2):
                text = parts[i + 1].strip().rstrip('.')
                if not text:
                    continue
                # 첫 문장이나 키워드를 제목으로 추출
                title_match = re.match(r'^(.+?)[.:\-—]', text)
                title = title_match.group(1).strip() if title_match else text[:70]
                items.append({"id": len(items), "title": title, "text": text, "status": "pending"})

    # 4) fallback
    if not items:
        items.append({"id": 0, "title": "전체 결론", "text": conclusion, "status": "pending"})

    return items


def finalize_team_notes(discuss: dict):
    """승인된 항목만 team_notes.md에 덮어쓰기 + pending_impl.json 생성."""
    import datetime
    items = discuss.get("conclusion_items", [])
    approved = [i for i in items if i["status"] == "approved"]
    topic = discuss.get("topic", "")
    today = datetime.datetime.now().strftime("%Y-%m-%d")

    content = TEAM_NOTES_HEADER
    if approved:
        content += f"\n## {topic}\n> 결정일: {today}\n\n"
        for item in approved:
            content += f"### {item['title']}\n{item['text']}\n\n"
        content += "---\n"

    TEAM_NOTES_PATH.write_text(content, encoding="utf-8")

    # 구현 대기 파일 생성 → UserPromptSubmit 훅이 감지해 Claude에 주입
    if approved:
        pending = {
            "status": "pending",
            "topic": topic,
            "approved_at": datetime.datetime.now().isoformat(),
            "items": approved,
        }
        PENDING_IMPL_PATH.write_text(
            json.dumps(pending, ensure_ascii=False, indent=2), encoding="utf-8"
        )


def _lookup_tc_title(nodeid: str, group: str) -> str:
    """nodeid → testcases/{group}/tc_*.md 에서 한글 제목 반환. 없으면 빈 문자열."""
    import re as _re
    parts = nodeid.split("/")
    py_file = parts[-1].split("::")[0]
    m = _re.match(r"(tc_(?:[A-Za-z]+_)?\d+)_", py_file)
    if not m:
        return ""
    tc_prefix = m.group(1)
    tc_dir = TESTCASES_DIR / group
    if not tc_dir.exists():
        return ""
    matches = sorted(tc_dir.glob(f"{tc_prefix}_*.md"))
    if not matches:
        return ""
    try:
        text = matches[0].read_text(encoding="utf-8")
    except OSError:
        return ""
    title_m = _re.search(r"^#\s+(.+)$", text, _re.MULTILINE)
    return title_m.group(1).strip() if title_m else ""


def _enrich_group_results(data: dict) -> dict:
    """execution_result.group_results 각 테스트에 한글 title 필드 추가."""
    exec_result = data.get("execution_result")
    if not exec_result or not isinstance(exec_result, dict):
        return data
    group_results = exec_result.get("group_results", {})
    if not group_results:
        return data
    for group, gdata in group_results.items():
        if not isinstance(gdata, dict):
            continue
        for test in gdata.get("tests", []):
            if not isinstance(test, dict) or test.get("title"):
                continue
            nodeid = test.get("nodeid", "")
            test["title"] = _lookup_tc_title(nodeid, group)
    return data


def build_pipeline_state() -> dict:
    """단일 파이프라인 state/pipeline.json 반환 (group_results에 한글 title 추가)."""
    data = load_json(STATE_PATH) or {}
    return _enrich_group_results(data)


def build_batch_state() -> dict:
    """병렬 파이프라인 상태 + tests/generated/ 파일 목록 반환."""
    parallel = load_json(PARALLEL_STATE_PATH) or {}
    generated_files = []
    if GENERATED_DIR.exists():
        for group_dir in sorted(GENERATED_DIR.iterdir()):
            if group_dir.is_dir() and not group_dir.name.startswith("."):
                for f in sorted(group_dir.glob("*.py")):
                    if f.name not in ("conftest.py", "__init__.py"):
                        generated_files.append({
                            "group": group_dir.name,
                            "file": f.name,
                            "path": str(f.relative_to(PROJECT_ROOT)),
                            "size": f.stat().st_size,
                        })
    # completed_count: parallel.json 값이 아닌 실제 생성 파일 수로 보정.
    # subagent 완료 후 parallel.json이 갱신되기 전에도 정확한 진행률을 표시하기 위함.
    if parallel.get("status") in (ParallelStatus.READY, "generating",  # P82: 상수 교체 ("generating"은 UI 파생 상태)
                                    ParallelStatus.TESTING, ParallelStatus.DONE):
        parallel = {**parallel, "completed_count": len(generated_files)}
    _enrich_group_results(parallel)
    return {"parallel_state": parallel, "generated_files": generated_files}


def build_pipeline_registry() -> dict:
    """프론트엔드용 파이프라인 레지스트리 상수 (P45).

    /api/pipeline_registry GET 엔드포인트가 반환하는 데이터.
    _pipeline_registry.py가 단일 소스 — 이 함수가 프론트 표현 형식으로 변환.
    constants.js가 이 값을 fetch해 PIPELINE_STEPS / STEP_LABELS 등 전역 변수를 갱신.
    """
    # 단일 파이프라인 스텝바 순서 (heal/timeout은 표시 이탈 상태이므로 제외)
    _terminal_excl = {Step.HEAL_NEEDED, Step.HEAL_FAILED, Step.TIMEOUT}
    pipeline_steps = [s.step for s in PIPELINE_STEP_DEFS if s.step not in _terminal_excl]

    # 모든 step 라벨 (heal 포함 — STEP_LABELS 전체 대체용)
    # M-2(P134): last-wins dict comprehension → STEP_DEF_BY_NAME 사용 (first-wins, STEP_DEF_BY_NAME과 동일 동작)
    step_labels: dict[str, str] = {k: v.label for k, v in STEP_DEF_BY_NAME.items()}
    # 구 step 값 호환 맵 — P58: _pipeline_registry.STEP_COMPAT이 단일 소스
    step_compat = STEP_COMPAT
    # compat step에도 라벨 추가 (STEP_LABELS[compat_step] 조회 지원)
    for alias, canonical in step_compat.items():
        step_labels.setdefault(alias, step_labels.get(canonical, alias))

    # 병렬 파이프라인 스텝바 순서
    # "generating"은 레지스트리 미등록 UI 파생 상태 (ready + files>0 조건)
    parallel_steps = [
        ParallelStatus.INIT,
        ParallelStatus.ANALYZING,
        ParallelStatus.READY,
        "generating",          # UI 파생 상태: parallel.js가 ready에서 추론
        ParallelStatus.TESTING,
        ParallelStatus.DONE,
    ]
    # P58: _pipeline_registry.PARALLEL_STEP_LABELS이 단일 소스
    parallel_step_labels = PARALLEL_STEP_LABELS

    return {
        "pipeline": {
            "steps":       pipeline_steps,
            "step_labels": step_labels,
            "step_compat": step_compat,
        },
        "parallel": {
            "steps":       parallel_steps,
            "step_labels": parallel_step_labels,
        },
    }


PAGES_JSON = PROJECT_ROOT / "config" / "pages.json"
TESTCASES_DIR = PROJECT_ROOT / "testcases"


def list_pages() -> dict:
    """config/pages.json 반환 (_comment 등 메타 키 제외)."""
    raw = load_json(PAGES_JSON) or {}
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def list_testcase_groups() -> list:
    """testcases/ 하위 폴더별 케이스 파일 목록."""
    if not TESTCASES_DIR.exists():
        return []
    groups = []
    for d in sorted(TESTCASES_DIR.iterdir()):
        if not d.is_dir() or d.name.startswith("."):
            continue
        cases = sorted([f.name for f in d.glob("tc_*.md")])
        groups.append({"name": d.name, "cases": cases, "count": len(cases)})
    return groups


def _natural_sort_key(name: str) -> list:
    """숫자 부분을 정수로 변환해 자연 정렬 키 반환 (tc_9 < tc_10 < tc_11 보장)."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r"(\d+)", name)]


def list_generated_groups() -> list:
    """tests/generated/ 하위 그룹별 테스트 파일 목록 반환.
    testcases/ 의 .md 파일 기준으로 유효한 파일만 집계 (잔여 파일 제외).
    """
    if not GENERATED_DIR.exists():
        return []
    groups = []
    for d in sorted(GENERATED_DIR.iterdir(), key=lambda p: _natural_sort_key(p.name)):
        if not d.is_dir() or d.name.startswith((".", "_")):
            continue
        all_py = sorted([
            f.name for f in d.glob("tc_*.py")
        ], key=_natural_sort_key)
        # testcases/{group}/tc_*.md 기준으로 유효 파일 집합 산출 (번호 prefix로 매칭)
        tc_dir = TESTCASES_DIR / d.name
        if tc_dir.exists():
            import re as _re
            def _tc_key(name):
                # tc_01_… 또는 tc_CL_01_… 형식 모두 지원 (영문 접두어 선택적)
                m = _re.match(r'^(tc_(?:[A-Za-z]+_)?\d+)_', name)
                return m.group(1) if m else None
            valid_keys = {_tc_key(f.name) for f in tc_dir.glob("tc_*.md")} - {None}
            files = [f for f in all_py if _tc_key(f) in valid_keys]
            stale_count = len(all_py) - len(files)
        else:
            files = all_py
            stale_count = 0
        if files:
            entry = {
                "name": d.name,
                "file_count": len(files),
                "files": files,
            }
            if stale_count > 0:
                entry["stale_count"] = stale_count
            groups.append(entry)
    return groups


MAX_REPORTS = 200


def list_reports() -> list:
    """tests/reports/ 의 HTML 파일 목록 (최신순, 최대 MAX_REPORTS개)."""
    if not REPORTS_DIR.exists():
        return []
    reports = []
    for f in sorted(REPORTS_DIR.glob("*.html"), key=lambda p: p.stat().st_mtime, reverse=True)[:MAX_REPORTS]:
        reports.append({
            "name": f.name,
            "modified_at": datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
            "size": f.stat().st_size,
        })
    return reports


def build_dialogs() -> dict:
    """팀 토론 대화 payload 반환 (dialog.json은 팀 토론 전용)."""
    full_dialog = load_json(DIALOG_PATH) or {"sessions": []}
    discuss_state = load_json(DISCUSS_PATH) or {}

    # step=discussed 이고 conclusion_items 없으면 메모리에서만 파싱 (P56: GET에서 write 금지)
    # 파싱 결과는 이 호출의 반환값에만 포함되고 파일에는 기록하지 않는다.
    # 영속화가 필요하면 별도 POST 엔드포인트를 사용한다.
    if (discuss_state.get("step") == "discussed"
            and discuss_state.get("conclusion")
            and not discuss_state.get("conclusion_items")):
        discuss_state = {**discuss_state,
                         "conclusion_items": parse_conclusion_items(discuss_state["conclusion"])}

    all_sessions = full_dialog.get("sessions", [])
    team_sessions = [s for s in all_sessions if s.get("stage") == "team_discussion"]

    # discuss_state의 conclusion_items와 status를 topic이 일치하는 세션에 주입
    if discuss_state.get("topic"):
        for ts in team_sessions:
            if ts.get("topic") == discuss_state["topic"]:
                if discuss_state.get("conclusion_items"):
                    ts["conclusion_items"] = discuss_state["conclusion_items"]
                if discuss_state.get("step"):
                    ts["status"] = discuss_state["step"]
                break

    return {
        "team_sessions": team_sessions,
        "discuss_state": discuss_state,
    }


# ── Excel Import 유틸 ─────────────────────────────────────────
def _list_import_files() -> list:
    """import/ 폴더의 .xlsx 파일 목록."""
    if not IMPORT_DIR.exists():
        return []
    return sorted([f.name for f in IMPORT_DIR.glob("*.xlsx")])


def _detect_header_row(ws) -> int | None:
    """'Test\\nScenario ID' 패턴이 있는 헤더 행 번호 반환."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        for cell in row:
            if cell.value and "Scenario ID" in str(cell.value):
                return i
    return None


def _list_excel_sheets(filepath: Path) -> list:
    """엑셀 파일의 시트별 테스트케이스 수 반환."""
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = _detect_header_row(ws)
        if header_row is None:
            continue
        count = 0
        for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
            cell_b = row[1].value if len(row) > 1 else None
            if cell_b and "_" in str(cell_b):
                count += 1
        if count > 0:
            sheets.append({"name": name, "count": count})
    wb.close()
    return sheets


def _parse_excel_sheet(wb, sheet_name: str) -> list:
    """엑셀 시트에서 테스트케이스 목록 추출."""
    ws = wb[sheet_name]
    header_row = _detect_header_row(ws)
    if header_row is None:
        return []

    last_main = ""
    last_sub = ""
    cases = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=False):
        tc_id = row[1].value if len(row) > 1 else None
        if not tc_id or "_" not in str(tc_id):
            continue
        main = str(row[2].value).strip() if len(row) > 2 and row[2].value else ""
        sub = str(row[3].value).strip() if len(row) > 3 and row[3].value else ""
        detail = str(row[4].value).strip() if len(row) > 4 and row[4].value else ""
        summary = str(row[5].value).strip() if len(row) > 5 and row[5].value else ""
        precond = str(row[6].value).strip() if len(row) > 6 and row[6].value else ""
        steps = str(row[7].value).strip() if len(row) > 7 and row[7].value else ""
        expected = str(row[8].value).strip() if len(row) > 8 and row[8].value else ""
        level = str(row[9].value).strip() if len(row) > 9 and row[9].value else ""
        if main:
            last_main = main
        else:
            main = last_main
        if sub:
            last_sub = sub
        else:
            sub = last_sub
        cases.append({
            "main": main, "sub": sub, "detail": detail,
            "summary": summary, "precondition": precond,
            "steps": steps, "expected": expected, "level": level,
        })
    return cases


def _level_to_priority(level: str) -> str:
    level = level.strip()
    if level in ("BAT", "Level 1"):
        return "high"
    elif level == "Level 2":
        return "medium"
    return "low"


def _to_slug(text: str) -> str:
    """TC Summary → 파일명 슬러그."""
    if not text:
        return "unnamed"
    text = text.replace("\n", " ").strip()
    text = re.sub(r'[/\\:*?"<>|.\[\]()>{},]', '', text)
    text = re.sub(r'\s+', '_', text)
    text = re.sub(r'_+', '_', text).strip('_')
    return text[:60]


def _write_tc_files(cases: list, output_dir: Path) -> int:
    """케이스 목록을 tc_*.md 파일로 생성. 생성 건수 반환."""
    output_dir.mkdir(parents=True, exist_ok=True)
    for i, c in enumerate(cases):
        num = str(i + 1).zfill(3)
        slug = _to_slug(c["summary"])
        filepath = output_dir / f"tc_{num}_{slug}.md"
        priority = _level_to_priority(c.get("level", ""))
        tags = []
        if c["main"]:
            clean = re.sub(r'\(.*?\)', '', c["main"]).strip().replace('\n', '')
            if clean:
                tags.append(clean)
        if c["sub"]:
            tags.append(c["sub"].replace('\n', ''))
        if not tags:
            tags = ["general"]
        tags_str = ", ".join(tags)
        steps_lines = [s.strip() for s in c["steps"].split("\n")
                       if s.strip() and not s.strip().startswith("0.")]
        steps_text = "\n".join(steps_lines) if steps_lines else "1. (스텝 미기재)"
        exp_lines = []
        for e in c["expected"].split("\n"):
            e = e.strip()
            if e:
                if not e.startswith("-") and not e.startswith("*"):
                    e = f"- {e}"
                exp_lines.append(e)
        expected_text = "\n".join(exp_lines) if exp_lines else "- (기대결과 미기재)"
        pre_lines = [p.strip() for p in c["precondition"].split("\n") if p.strip()]
        precond_text = "\n".join(pre_lines) if pre_lines else "- 없음"
        title = c["summary"].replace("\n", " ").strip()
        content = (
            f"---\nid: tc_{num}\ndata_key: null\npriority: {priority}\n"
            f"tags: [{tags_str}]\ntype: structured\n---\n"
            f"# {title}\n\n## Precondition\n{precond_text}\n\n"
            f"## Steps\n{steps_text}\n\n## Expected\n{expected_text}\n"
        )
        filepath.write_text(content, encoding="utf-8")
    return len(cases)


def _read_body(handler) -> dict:
    """요청 바디를 JSON으로 파싱해 반환. 바디 없으면 빈 dict."""
    length = int(handler.headers.get("Content-Length", 0))
    return json.loads(handler.rfile.read(length).decode("utf-8")) if length else {}


def _read_profiles_locked(path: Path) -> dict:
    from _paths import _file_lock
    from _import_commit import ImportRunError
    path.parent.mkdir(parents=True, exist_ok=True)
    with _file_lock(path.with_suffix(".lock"), path):
        if not path.exists():
            return {"profiles": []}
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ImportRunError("프로필 저장소를 읽을 수 없습니다", "PROFILE_STORE_ERROR") from exc
        if not isinstance(data, dict) or not isinstance(data.get("profiles", []), list):
            raise ImportRunError("프로필 저장소 형식이 잘못되었습니다", "PROFILE_STORE_ERROR")
        return data


def _update_profiles_locked(path: Path, mutate) -> dict:
    import tempfile
    from _paths import _file_lock
    from _import_commit import ImportRunError
    path.parent.mkdir(parents=True, exist_ok=True)
    with _file_lock(path.with_suffix(".lock"), path):
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError) as exc:
                raise ImportRunError("프로필 저장소를 읽을 수 없습니다", "PROFILE_STORE_ERROR") from exc
            if not isinstance(data, dict) or not isinstance(data.get("profiles", []), list):
                raise ImportRunError("프로필 저장소 형식이 잘못되었습니다", "PROFILE_STORE_ERROR")
        else:
            data = {"profiles": []}
        result = mutate(data)
        fd, tmp = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as stream:
                json.dump(data, stream, ensure_ascii=False, indent=2)
            Path(tmp).replace(path)
        except Exception:
            Path(tmp).unlink(missing_ok=True)
            raise
        return result


class DashboardHandler(BaseHTTPRequestHandler):

    # ── Route 딕셔너리 ────────────────────────────────────────────
    GET_ROUTES = {
        "/api/dialogs":          "_get_dialogs",
        "/api/events":           "_get_events",
        "/api/dialog":           "_get_dialog",
        "/api/state":            "_get_state",
        "/api/pages":            "_get_pages",
        "/api/pipeline_state":   "_get_pipeline_state",
        "/api/batch_state":      "_get_batch_state",
        "/api/quick_state":      "_get_quick_state",
        "/api/run_history":      "_get_run_history",
        "/api/heal_stats":       "_get_heal_stats",
        "/api/pipeline_registry": "_get_pipeline_registry",
        "/api/flaky_tests":      "_get_flaky_tests",
        "/api/generated_groups": "_get_generated_groups",
        "/api/reports":          "_get_reports",
        "/api/testcase":         "_get_testcase",
        "/api/import/files":     "_get_import_files",
        "/api/import/sheets":    "_get_import_sheets",
        "/api/import/profiles":     "_get_import_profiles_v2",
        "/api/import/preview/csv":  "_get_import_preview_csv",
    }

    POST_ROUTES = {
        "/api/pages/add":          "_post_pages_add",
        "/api/pages/update":       "_post_pages_update",
        "/api/pages/delete":       "_post_pages_delete",
        "/api/reset":              "_post_reset",
        "/api/reset/all":          "_post_reset_all",
        "/api/discuss/start":      "_post_discuss_start",
        "/api/discuss/vote_item":  "_post_discuss_vote_item",
        "/api/discuss/reject":     "_post_discuss_reject",
        "/api/run_qa":             "_post_run_qa",
        "/api/run_qa_parallel":    "_post_run_qa_parallel",
        "/api/run_log":            "_post_run_log",
        "/api/pipeline/reset":     "_post_pipeline_reset",
        "/api/parallel/reset":     "_post_parallel_reset",
        "/api/quick/reset":        "_post_quick_reset",
        "/api/run_merge":          "_post_run_merge",
        "/api/merge_log":          "_post_merge_log",
        "/api/run_quick":          "_post_run_quick",
        "/api/import/convert":          "_post_import_convert",
        "/api/import/preview":          "_post_import_preview_v2",
        "/api/import/profiles":         "_post_import_profiles_v2",
        "/api/import/profiles/update":  "_post_import_profiles_update",
        "/api/import/profiles/delete":  "_post_import_profiles_delete",
        "/api/import/commit":           "_post_import_commit_v2",
        "/api/import/rollback":         "_post_import_rollback_v2",
        "/api/heal_stats/reset":        "_post_heal_stats_reset",
        "/api/run_history/reset":  "_post_run_history_reset",
        "/api/discuss/reset":      "_post_discuss_reset",
    }

    # ── Dispatchers ───────────────────────────────────────────────
    def do_GET(self):
        path = self.path.split("?")[0]

        # index
        if path in ("/", "/index.html", "/import-studio"):
            self._serve_file(HERE / "index.html", "text/html; charset=utf-8")
            return

        # exact-match routes
        if path in self.GET_ROUTES:
            getattr(self, self.GET_ROUTES[path])()
            return

        if path.startswith("/api/import/runs/"):
            self._get_import_run_v2(path)
            return

        # prefix routes
        if path.startswith("/reports/"):
            self._get_report_file(path)
            return

        if path.startswith("/screenshots/"):
            self._get_artifact_file(path, "/screenshots/", SCREENSHOTS_DIR, "image/png")
            return

        if path.startswith("/videos/"):
            self._get_artifact_file(path, "/videos/", VIDEOS_DIR, "video/mp4")
            return

        if path.startswith("/static/"):
            self._get_static_file(path)
            return

        self.send_response(404)
        self.end_headers()

    def _check_csrf_origin(self) -> bool:
        """브라우저發 크로스사이트 POST 및 DNS rebinding을 차단한다 (P49).

        1. Host 헤더 검증 (DNS rebinding 방어):
           Host가 ALLOWED_HOSTS 목록에 없으면 거부.
           Origin/Referer 유무와 무관하게 항상 적용.

        2. Origin/Referer 검증 (CSRF 방어):
           Origin이 있으면 ALLOWED_ORIGIN과 비교.
           Referer가 있으면 스킴+호스트만 비교.
           둘 다 없는 요청(curl 등 로컬 CLI)은 Host 검증 통과 시 허용.
        """
        from urllib.parse import urlparse

        # ── (1) Host 헤더 검증 ─────────────────────────────────
        host = self.headers.get("Host", "")
        # Host가 아예 없으면(HTTP/1.0) localhost로 간주해 허용
        if host and host not in ALLOWED_HOSTS:
            return False

        # ── (2) Origin / Referer 검증 ──────────────────────────
        origin = self.headers.get("Origin")
        if origin is None:
            referer = self.headers.get("Referer")
            if referer is None:
                # 헤더 없음 = curl 등 로컬 CLI → Host 검증 통과했으면 허용
                return True
            # Referer는 경로까지 포함하므로 스킴+호스트만 비교
            parsed = urlparse(referer)
            origin = f"{parsed.scheme}://{parsed.netloc}" if parsed.scheme else referer
        return origin == ALLOWED_ORIGIN

    def _check_remote_allowlist(self, path: str) -> bool:
        """Apply the remote-mode exact/literal-prefix mutation allowlist."""
        if not REMOTE_MODE:
            return True

        is_run = path.startswith("/api/run_")
        is_reset = (
            path == "/api/reset"
            or path.startswith("/api/reset/")
            or path.endswith("/reset")
        )
        if not (is_run or is_reset):
            return True

        for pattern in REMOTE_API_ALLOWLIST:
            if pattern.endswith("*"):
                if path.startswith(pattern[:-1]):
                    return True
            elif path == pattern:
                return True
        return False

    def do_POST(self):
        path = self.path.split("?")[0]
        if not self._check_csrf_origin():
            self.send_response(403)
            self.end_headers()
            return
        if not self._check_remote_allowlist(path):
            content = json.dumps(
                {
                    "ok": False,
                    "error": (
                        "remote mode blocks this endpoint unless allowlisted: "
                        f"{path}"
                    ),
                },
                ensure_ascii=False,
            ).encode("utf-8")
            self.send_response(403)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return
        try:
            if path.startswith("/api/import/runs/") and path.endswith("/rollback"):
                self._post_import_run_rollback_v2(path)
                return
            if path in self.POST_ROUTES:
                getattr(self, self.POST_ROUTES[path])()
            else:
                self.send_response(404)
                self.end_headers()
        except Exception as e:
            import traceback
            traceback.print_exc()
            msg = json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False).encode("utf-8")
            try:
                msg = json.dumps({"ok": False, "error": "internal server error",
                                  "code": "INTERNAL_ERROR"}, ensure_ascii=False).encode("utf-8")
                self._serve_bytes(msg, "application/json; charset=utf-8", status=500)
            except Exception:
                pass

    def do_DELETE(self):
        """REST-compatible mapping profile deletion."""
        path = self.path.split("?")[0]
        if not self._check_csrf_origin():
            self.send_response(403)
            self.end_headers()
            return
        prefix = "/api/import/profiles/"
        if path.startswith(prefix):
            self._delete_import_profile_v2(path[len(prefix):])
            return
        self.send_response(404)
        self.end_headers()

    def do_PUT(self):
        """REST-compatible mapping profile update."""
        path = self.path.split("?")[0]
        if not self._check_csrf_origin():
            self.send_response(403)
            self.end_headers()
            return
        prefix = "/api/import/profiles/"
        if path.startswith(prefix):
            self._put_import_profile_v2(path[len(prefix):])
            return
        self.send_response(404)
        self.end_headers()

    # ── GET handlers ─────────────────────────────────────────────
    def _get_dialogs(self):
        payload = build_dialogs()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_events(self):
        self._serve_sse()

    def _get_dialog(self):
        self._serve_json(DIALOG_PATH)

    def _get_state(self):
        self._serve_json(STATE_PATH)

    def _get_pages(self):
        pages = list_pages()
        # project 목록 추출 (중복 제거, 알파벳 정렬 — "기본"은 UI 개념이므로 미포함)
        project_set: set[str] = set()
        for k, v in pages.items():
            if k == "_comment":
                continue
            if isinstance(v, dict) and v.get("project"):
                project_set.add(v["project"])
        project_list = sorted(project_set)
        payload = {"pages": pages, "groups": list_testcase_groups(), "projects": project_list}
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_pipeline_state(self):
        payload = build_pipeline_state()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_batch_state(self):
        payload = build_batch_state()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_quick_state(self):
        payload = load_json(QUICK_STATE_PATH) or {}
        payload = _enrich_group_results(payload)
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_run_history(self):
        payload = load_json(RUN_HISTORY_PATH) or []
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_flaky_tests(self):
        payload = load_json(FLAKY_TESTS_PATH) or {"flaky": []}
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_heal_stats(self):
        payload = load_json(HEAL_STATS_PATH) or {}
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_pipeline_registry(self):
        """P45: _pipeline_registry.py 상수를 JSON으로 노출. constants.js가 fetch."""
        payload = build_pipeline_registry()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_generated_groups(self):
        payload = list_generated_groups()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_reports(self):
        payload = list_reports()
        self._serve_bytes(
            json.dumps(payload, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_testcase(self):
        import re as _re
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        nodeid = qs.get("nodeid", [""])[0]
        if not nodeid:
            self._serve_bytes(
                b'{"ok":false,"error":"nodeid parameter required"}',
                "application/json; charset=utf-8")
            return
        # nodeid 예: tests/heroku/tc_01_login.py::test_login
        #           tests/generated/directcloud/tc_01_login_success_login_success.py::test_...
        parts = nodeid.split("/")
        py_file = parts[-1].split("::")[0]  # tc_01_*.py
        # group = 마지막 디렉토리 (generated 건너뜀)
        group = parts[-2] if len(parts) >= 2 else ""
        if group == "generated" and len(parts) >= 3:
            group = parts[-2]  # generated/{group}/file 구조에선 이미 parts[-2]가 group
        if not group or not py_file or ".." in group:
            self._serve_bytes(
                b'{"ok":false,"error":"invalid nodeid"}',
                "application/json; charset=utf-8")
            return
        # tc 번호 추출: tc_01_ / tc_CL_01_ → testcases/{group}/tc_*_.md 검색
        m = _re.match(r"(tc_(?:[A-Za-z]+_)?\d+)_", py_file)
        tc_prefix = m.group(1) if m else None
        tc_dir = TESTCASES_DIR / group
        fpath = None
        if tc_prefix and tc_dir.exists():
            matches = sorted(tc_dir.glob(f"{tc_prefix}_*.md"))
            if matches:
                fpath = matches[0]
        if fpath is None or not fpath.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"tc file not found for {py_file}"}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        content = fpath.read_text(encoding="utf-8")
        self._serve_bytes(
            json.dumps({"ok": True, "content": content, "file": fpath.name}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _get_import_files(self):
        """GET /api/import/files — 메타데이터 포함 파일 목록 (S2 확장)."""
        if not IMPORT_DIR.exists():
            # import/ 폴더가 없으면 빈 배열 반환 (FE 오류 방지)
            self._serve_bytes(
                json.dumps({"ok": True, "files": []}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8",
            )
            return

        import hashlib as _hashlib
        from _excel_import import get_file_metadata  # type: ignore[import]

        files = []
        for f in sorted(IMPORT_DIR.glob("*.xlsx")):
            file_id = _hashlib.sha256(f.name.encode("utf-8")).hexdigest()[:8]
            try:
                meta = get_file_metadata(f)
            except Exception as exc:
                # 파싱 불가 파일도 목록에는 포함 (sheets 없이)
                meta = {"sheets": [], "size": f.stat().st_size,
                        "modified": "", "error": str(exc)}
            files.append({
                "id":       file_id,
                "name":     f.name,
                "size":     meta.get("size", 0),
                "modified": meta.get("modified", ""),
                "sheets":   meta.get("sheets", []),
                "error":    meta.get("error"),
            })
        self._serve_bytes(
            json.dumps({"ok": True, "files": files}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _get_import_sheets(self):
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        fname = qs.get("file", [""])[0]
        if not is_safe_filename(fname):
            self._serve_bytes(
                b'{"ok":false,"error":"file parameter required"}',
                "application/json; charset=utf-8")
            return
        fpath = IMPORT_DIR / fname
        if not fpath.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"{fname} not found"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        try:
            sheets = _list_excel_sheets(fpath)
            self._serve_bytes(
                json.dumps({"ok": True, "sheets": sheets},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
        except Exception as e:
            self._serve_bytes(
                json.dumps({"ok": False, "error": str(e)},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")

    def _get_report_file(self, path: str):
        fname = path[len("/reports/"):]
        if not is_safe_filename(fname):
            self.send_response(403)
            self.end_headers()
            return
        fpath = REPORTS_DIR / fname
        if fpath.exists() and fpath.suffix == ".html":
            content = fpath.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(content)))
            self.send_header("X-XSS-Protection", "0")
            self.send_header("Content-Security-Policy",
                             "default-src 'self' https:; "
                             "script-src 'unsafe-inline'; "
                             "style-src 'unsafe-inline' https:; "
                             "font-src 'self' https: data:; "
                             "img-src 'self' data: blob:; "
                             "media-src 'self' blob:")
            self.end_headers()
            self.wfile.write(content)
        else:
            self.send_response(404)
            self.end_headers()

    def _get_artifact_file(self, path: str, prefix: str, base_dir: Path, content_type: str):
        """screenshots / videos 등 아티팩트 파일 서빙."""
        fname = path[len(prefix):]
        if not is_safe_filename(fname):
            self.send_response(403)
            self.end_headers()
            return
        fpath = base_dir / fname
        if fpath.exists() and fpath.is_file():
            self._serve_file(fpath, content_type)
        else:
            self.send_response(404)
            self.end_headers()

    def _get_static_file(self, path: str):
        rel = path[len("/static/"):]
        # P66: ".." 단독 검사는 절대경로 주입을 막지 못함 → resolve 후 봉쇄
        _static_root = (HERE / "static").resolve()
        fpath = (_static_root / rel).resolve()
        if not fpath.is_relative_to(_static_root):
            self.send_response(403)
            self.end_headers()
            return
        if fpath.exists() and fpath.is_file():
            ext = fpath.suffix.lower()
            mime_map = {
                ".css": "text/css; charset=utf-8",
                ".js": "application/javascript; charset=utf-8",
                ".png": "image/png",
                ".svg": "image/svg+xml",
            }
            content_type = mime_map.get(ext, "application/octet-stream")
            self._serve_file(fpath, content_type)
        else:
            self.send_response(404)
            self.end_headers()

    # ── POST handlers ─────────────────────────────────────────────
    def _post_pages_add(self):
        body = _read_body(self)
        group   = body.get("group", "").strip()
        url     = body.get("url", "").strip()
        notes   = body.get("notes", "").strip()
        spa     = bool(body.get("spa", False))
        project = body.get("project", "").strip()

        if not group or not url:
            self._serve_bytes(b'{"ok":false,"error":"group and url required"}', "application/json; charset=utf-8"); return
        if not is_valid_group_name(group) or group.startswith("_"):
            self._serve_bytes(b'{"ok":false,"error":"invalid group name"}', "application/json; charset=utf-8"); return
        if not is_valid_url(url):
            self._serve_bytes(b'{"ok":false,"error":"url must start with http:// or https://"}', "application/json; charset=utf-8"); return
        if project and not is_valid_group_name(project):
            self._serve_bytes(b'{"ok":false,"error":"project name: alphanumeric/underscore/hyphen only"}', "application/json; charset=utf-8"); return

        # pages.json 파싱 실패 시 설정 소실 방지
        if PAGES_JSON.exists():
            raw = load_json(PAGES_JSON)
            if raw is None:
                self._serve_bytes(
                    b'{"ok":false,"error":"pages.json \xed\x8c\x8c\xec\x8b\xb1 \xec\x8b\xa4\xed\x8c\xa8 \xe2\x80\x94 \xec\x88\x98\xeb\x8f\x99 \xed\x99\x95\xec\x9d\xb8 \xed\x95\x84\xec\x9a\x94"}',
                    "application/json; charset=utf-8"); return

        # RMW 원자적 처리 (락 보유 중 읽기+중복 체크+쓰기)
        error_msg = None
        def _add_mutator(cur: dict) -> dict:
            nonlocal error_msg
            if group in cur:
                error_msg = f"'{group}' 그룹이 이미 존재합니다"
                return cur
            entry: dict = {"url": url, "spa": spa, "preconditions": [], "notes": notes}
            if project:
                entry["project"] = project
            cur[group] = entry
            return cur

        _safe_update_json(PAGES_JSON, _add_mutator)
        if error_msg:
            self._serve_bytes(
                json.dumps({"ok": False, "error": error_msg}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"); return

        (TESTCASES_DIR / group).mkdir(parents=True, exist_ok=True)
        print(f"[Dashboard] 페이지 추가: {group} → {url}")
        self._serve_bytes(
            json.dumps({"ok": True, "group": group}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8")

    def _post_pages_update(self):
        body    = _read_body(self)
        group   = body.get("group", "").strip()
        url     = body.get("url", "").strip()
        notes   = body.get("notes", "").strip()
        spa     = bool(body.get("spa", False))
        project = body.get("project", "").strip()

        if not group or not url:
            self._serve_bytes(b'{"ok":false,"error":"group and url required"}', "application/json; charset=utf-8"); return
        if not is_valid_group_name(group) or group.startswith("_"):
            self._serve_bytes(b'{"ok":false,"error":"invalid group name"}', "application/json; charset=utf-8"); return
        if not is_valid_url(url):
            self._serve_bytes(b'{"ok":false,"error":"url must start with http:// or https://"}', "application/json; charset=utf-8"); return
        if project and not is_valid_group_name(project):
            self._serve_bytes(b'{"ok":false,"error":"project name: alphanumeric/underscore/hyphen only"}', "application/json; charset=utf-8"); return

        # pages.json 파싱 실패 시 설정 소실 방지 (P54)
        if PAGES_JSON.exists():
            raw = load_json(PAGES_JSON)
            if raw is None:
                self._serve_bytes(
                    b'{"ok":false,"error":"pages.json \xed\x8c\x8c\xec\x8b\xb1 \xec\x8b\xa4\xed\x8c\xa8 \xe2\x80\x94 \xec\x88\x98\xeb\x8f\x99 \xed\x99\x95\xec\x9d\xb8 \xed\x95\x84\xec\x9a\x94"}',
                    "application/json; charset=utf-8"); return

        error_msg = None
        def _update_mutator(cur: dict) -> dict:
            nonlocal error_msg
            if group not in cur:
                error_msg = f"'{group}' 없음"
                return cur
            existing = cur[group] if isinstance(cur[group], dict) else {}
            entry: dict = {
                "url": url,
                "spa": spa,
                "preconditions": existing.get("preconditions", []),
                "notes": notes,
            }
            if project:
                entry["project"] = project
            cur[group] = entry
            return cur

        _safe_update_json(PAGES_JSON, _update_mutator)
        if error_msg:
            self._serve_bytes(
                json.dumps({"ok": False, "error": error_msg}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"); return

        print(f"[Dashboard] 페이지 수정: {group} → {url}")
        self._serve_bytes(
            json.dumps({"ok": True, "group": group}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8")

    def _post_pages_delete(self):
        body = _read_body(self)
        group = body.get("group", "").strip()
        if not group or not is_valid_group_name(group) or group.startswith("_"):
            self._serve_bytes(b'{"ok":false,"error":"valid group required"}', "application/json; charset=utf-8"); return

        # pages.json 파싱 실패 시 설정 소실 방지 (P54)
        if PAGES_JSON.exists():
            raw = load_json(PAGES_JSON)
            if raw is None:
                self._serve_bytes(
                    b'{"ok":false,"error":"pages.json \xed\x8c\x8c\xec\x8b\xb1 \xec\x8b\xa4\xed\x8c\xa8 \xe2\x80\x94 \xec\x88\x98\xeb\x8f\x99 \xed\x99\x95\xec\x9d\xb8 \xed\x95\x84\xec\x9a\x94"}',
                    "application/json; charset=utf-8"); return

        # RMW 원자적 처리
        error_msg = None
        def _del_mutator(cur: dict) -> dict:
            nonlocal error_msg
            if group not in cur:
                error_msg = f"'{group}' 없음"
                return cur
            del cur[group]
            return cur

        _safe_update_json(PAGES_JSON, _del_mutator)
        if error_msg:
            self._serve_bytes(
                json.dumps({"ok": False, "error": error_msg}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"); return

        print(f"[Dashboard] 페이지 삭제: {group}")
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_reset(self):
        empty = {"pipeline_url": "", "started_at": "", "sessions": []}
        _safe_write_json(DIALOG_PATH, empty)
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_discuss_start(self):
        import datetime
        body = _read_body(self)
        topic = body.get("topic", "").strip()
        if not topic:
            self._serve_bytes(b'{"ok":false,"error":"topic required"}',
                              "application/json; charset=utf-8")
            return

        history = []
        if DISCUSS_PATH.exists():
            try:
                prev = json.loads(DISCUSS_PATH.read_text(encoding="utf-8"))
                history = prev.get("history", [])
                if prev.get("step") in ("approved", "rejected", "discussed"):
                    history.append({k: v for k, v in prev.items() if k != "history"})
            except Exception:
                pass

        discuss = {
            "topic": topic, "step": "pending", "conclusion": "",
            "rejection_reason": "", "rejection_count": 0,
            "created_at": datetime.datetime.now().isoformat(),
            "history": history,
        }
        _safe_write_json(DISCUSS_PATH, discuss)

        # Claude Code UserPromptSubmit 훅(check_pending_discuss.py)이
        # 다음 프롬프트 제출 시 자동으로 토론 시작을 Claude에게 주입한다.
        self._serve_bytes(b'{"ok":true}',
                          "application/json; charset=utf-8")

    def _post_discuss_vote_item(self):
        body = _read_body(self)
        item_id = int(body.get("item_id", -1))
        vote    = body.get("vote", "")  # "approve" | "reject"

        if not DISCUSS_PATH.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": "state/discuss.json 없음"}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return

        # P55: 비원자 read_text+write → _safe_update_json 원자적 RMW
        _result: dict = {}

        def _mutate_vote(s: dict) -> dict:
            items = [dict(i) for i in s.get("conclusion_items", [])]
            for item in items:
                if item["id"] == item_id:
                    item["status"] = "approved" if vote == "approve" else "rejected"
                    break
            s = {**s, "conclusion_items": items}
            all_voted = bool(items) and all(i["status"] != "pending" for i in items)
            if all_voted:
                finalize_team_notes(s)
                s["step"] = "approved"
            _result["all_voted"] = all_voted
            return s

        _safe_update_json(DISCUSS_PATH, _mutate_vote)
        self._serve_bytes(
            json.dumps({"ok": True, "all_voted": _result.get("all_voted", False)},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _post_discuss_reject(self):
        body = _read_body(self)
        reason = body.get("reason", "").strip()
        if not DISCUSS_PATH.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": "state/discuss.json 없음"}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        # P55: 비원자 read_text+write → _safe_update_json 원자적 RMW
        _safe_update_json(DISCUSS_PATH, lambda s: {
            **s,
            "step": "rejected",
            "rejection_reason": reason,
            "rejection_count": s.get("rejection_count", 0) + 1,
        })
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_run_qa(self):
        import subprocess as sp
        # P77: 중복 spawn 방지 — 이미 실행 중이면 거부
        if _is_script_running("run_qa"):
            self._serve_bytes(
                b'{"ok":false,"error":"run_qa already running"}',
                "application/json; charset=utf-8")
            return
        body = _read_body(self)
        url = body.get("url", "").strip()
        cases_dir = body.get("cases_dir", "").strip()  # e.g. "login"
        # 입력 검증: URL은 http(s)로 시작, cases_dir은 영숫자/언더스코어/하이픈만
        if url and not is_valid_url(url):
            self._serve_bytes(
                b'{"ok":false,"error":"url must start with http:// or https://"}',
                "application/json; charset=utf-8")
            return
        if cases_dir and not is_valid_group_name(cases_dir):
            self._serve_bytes(
                b'{"ok":false,"error":"invalid cases_dir format"}',
                "application/json; charset=utf-8")
            return
        if not url or not cases_dir:
            self._serve_bytes(
                b'{"ok":false,"error":"url and cases_dir required"}',
                "application/json; charset=utf-8")
            return
        cases_path = TESTCASES_DIR / cases_dir
        if not cases_path.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"testcases/{cases_dir} not found"}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        log_path = LOGS_DIR / "run_qa.txt"
        script = PROJECT_ROOT / "run_qa.py"
        log_file = open(log_path, "w", encoding="utf-8")
        proc = sp.Popen(
            [PYTHON_EXE, "-u", str(script),
             "--url", url, "--cases", str(cases_path)],
            cwd=str(PROJECT_ROOT),
            stdout=log_file, stderr=sp.STDOUT,
        )
        _register_spawned_proc(proc, tag="run_qa")  # P77: tag 등록
        # 자식 프로세스가 fd를 상속했으므로 부모에서 닫아도 안전
        log_file.close()
        print(f"[Dashboard] run_qa.py 실행 (PID: {proc.pid}, URL: {url}, cases: {cases_dir})")
        self._serve_bytes(
            json.dumps({"ok": True, "pid": proc.pid, "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _post_run_qa_parallel(self):
        import subprocess as sp
        # P77: 중복 spawn 방지
        if _is_script_running("run_qa_parallel"):
            self._serve_bytes(
                b'{"ok":false,"error":"run_qa_parallel already running"}',
                "application/json; charset=utf-8")
            return
        log_path = LOGS_DIR / "run_parallel.txt"
        script = PROJECT_ROOT / "run_qa_parallel.py"
        log_file = open(log_path, "w", encoding="utf-8")
        proc = sp.Popen(
            [PYTHON_EXE, "-u", str(script)],
            cwd=str(PROJECT_ROOT),
            stdout=log_file, stderr=sp.STDOUT,
        )
        _register_spawned_proc(proc, tag="run_qa_parallel")  # P77: tag 등록
        log_file.close()
        print(f"[Dashboard] run_qa_parallel.py 실행 (PID: {proc.pid})")
        self._serve_bytes(
            json.dumps({"ok": True, "pid": proc.pid, "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _post_run_log(self):
        body = _read_body(self)
        log_name = body.get("log", "run_qa.txt")
        if not is_safe_filename(log_name):
            self._serve_bytes(b'{"ok":false,"log":""}', "application/json; charset=utf-8")
            return
        log_path = LOGS_DIR / log_name
        if log_path.exists():
            content = log_path.read_text(encoding="utf-8", errors="replace")
            self._serve_bytes(
                json.dumps({"ok": True, "log": content}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        else:
            self._serve_bytes(b'{"ok":false,"log":""}', "application/json; charset=utf-8")

    def _post_reset_all(self):
        # dialog
        empty = {"pipeline_url": "", "started_at": "", "sessions": []}
        _safe_write_json(DIALOG_PATH, empty)
        # pipeline — 팩토리 함수로 단일화 (P39)
        # step/status를 Step.INIT/ParallelStatus.EMPTY로 되돌리는 리셋이라
        # FSM 전이 검증을 건너뛰는 reset_state()를 쓴다 — write_state()를 쓰면
        # 예: step="generated"에서 reset하면 VALID_TRANSITIONS에
        # "generated"→"init" 전이가 없어 ValueError로 리셋 자체가 실패한다.
        reset_state(STATE_PATH, make_initial_pipeline_state())
        # parallel
        reset_state(PARALLEL_STATE_PATH,
                    {"status": ParallelStatus.EMPTY, "total_count": 0, "targets": []})
        heal_ctx = PROJECT_ROOT / "state" / "heal_context.json"
        if heal_ctx.exists():
            heal_ctx.unlink()
        # quick
        if QUICK_STATE_PATH.exists():
            QUICK_STATE_PATH.unlink()
        # heal stats
        _safe_write_json(HEAL_STATS_PATH, {"version": 1, "patterns": {}})
        # run history
        _safe_write_json(RUN_HISTORY_PATH, [])
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_pipeline_reset(self):
        # 팩토리 함수로 단일화 (P39). FSM 검증 우회 이유는 _post_reset_all 참조.
        reset_state(STATE_PATH, make_initial_pipeline_state())
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_parallel_reset(self):
        init_state = {"status": ParallelStatus.EMPTY, "total_count": 0, "targets": []}
        reset_state(PARALLEL_STATE_PATH, init_state)
        # heal_context도 정리
        heal_ctx = PROJECT_ROOT / "state" / "heal_context.json"
        if heal_ctx.exists():
            heal_ctx.unlink()
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_quick_reset(self):
        import subprocess as sp
        body = _read_body(self)
        pid = body.get("pid") if body else None
        if pid:
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                self.send_response(400)
                self.end_headers()
                return
            # 이 서버가 직접 띄운 프로세스만 종료 대상으로 허용
            if not _is_spawned_pid(pid_int):
                self._serve_bytes(
                    json.dumps({"ok": False,
                                "error": f"이 서버가 생성한 프로세스가 아닙니다 (PID: {pid_int})"},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8", status=403)
                return
            try:
                with _SPAWNED_PIDS_LOCK:
                    tracked = _SPAWNED_PROCS.get(pid_int)
                if tracked is not None and tracked.poll() is None:
                    # P61: poll()으로 생존 확인 후 terminate
                    if sys.platform == "win32":
                        sp.run(["taskkill", "/F", "/T", "/PID", str(pid_int)],
                               capture_output=True, timeout=5)
                    else:
                        tracked.terminate()
            except Exception:
                pass
        if QUICK_STATE_PATH.exists():
            QUICK_STATE_PATH.unlink()
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_heal_stats_reset(self):
        init = {"version": 1, "patterns": {}}
        _safe_write_json(HEAL_STATS_PATH, init)
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_run_history_reset(self):
        _safe_write_json(RUN_HISTORY_PATH, [])
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_discuss_reset(self):
        # discuss.json 초기화 (topic/step/conclusion 등 모든 상태 제거)
        _safe_write_json(DISCUSS_PATH, {})
        # dialog.json의 team_discussion 세션도 제거
        dialog = load_json(DIALOG_PATH) or {"sessions": []}
        dialog["sessions"] = [
            s for s in dialog.get("sessions", [])
            if s.get("stage") != "team_discussion"
        ]
        _safe_write_json(DIALOG_PATH, dialog)
        self._serve_bytes(b'{"ok":true}', "application/json; charset=utf-8")

    def _post_run_merge(self):
        import subprocess as sp
        # P77: 중복 spawn 방지
        if _is_script_running("run_merge"):
            self._serve_bytes(
                b'{"ok":false,"error":"99_merge already running"}',
                "application/json; charset=utf-8")
            return
        merge_script = PROJECT_ROOT / "parallel" / "99_merge.py"
        if not merge_script.exists():
            self._serve_bytes(b'{"ok":false,"error":"99_merge.py not found"}',
                              "application/json; charset=utf-8")
            return
        # C-3(P99): heal_count 리셋은 "새 실행" 시작 시에만. 힐링 재실행(HEAL_NEEDED)에서
        # 리셋하면 MAX_HEAL 가드가 무력화되어 무한 루프가 발생한다.
        # M-4(P121): RESETTABLE_PARALLEL_STATUSES를 단일 소스에서 임포트 (3중 하드코딩 통합).
        if PARALLEL_STATE_PATH.exists():
            _cur_status = (load_json(PARALLEL_STATE_PATH) or {}).get("status", "")
            if _cur_status in RESETTABLE_PARALLEL_STATUSES:
                _safe_update_json(PARALLEL_STATE_PATH, lambda s: {**s, "heal_count": 0})

        # 로그 파일로 출력 저장
        log_path = LOGS_DIR / "merge.txt"
        log_file = open(log_path, "w", encoding="utf-8")
        proc = sp.Popen(
            [PYTHON_EXE, "-u", str(merge_script)],
            cwd=str(PROJECT_ROOT),
            stdout=log_file, stderr=sp.STDOUT,
        )
        _register_spawned_proc(proc, tag="run_merge")  # P77: tag 등록
        log_file.close()
        print(f"[Dashboard] 99_merge.py 실행 (PID: {proc.pid}, 로그: {log_path})")
        self._serve_bytes(
            json.dumps({"ok": True, "message": "99_merge.py started", "pid": proc.pid,
                         "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _post_merge_log(self):
        log_path = LOGS_DIR / "merge.txt"
        if log_path.exists():
            content = log_path.read_text(encoding="utf-8", errors="replace")
            self._serve_bytes(
                json.dumps({"ok": True, "log": content}, ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8"
            )
        else:
            self._serve_bytes(b'{"ok":false,"log":""}',
                              "application/json; charset=utf-8")

    def _post_run_quick(self):
        import subprocess as sp
        # P77: 중복 spawn 방지
        if _is_script_running("run_quick"):
            self._serve_bytes(
                b'{"ok":false,"error":"run_quick already running"}',
                "application/json; charset=utf-8")
            return
        body = _read_body(self)
        groups = body.get("groups", [])
        if not groups:
            self._serve_bytes(
                b'{"ok":false,"error":"groups required"}',
                "application/json; charset=utf-8")
            return
        # 그룹명 형식 검증 (경로 탈출 방지)
        invalid = [g for g in groups if not is_valid_group_name(g)]
        if invalid:
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"잘못된 그룹명: {', '.join(invalid)}"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        # 폴더 존재 검증
        missing = [g for g in groups if not (GENERATED_DIR / g).is_dir()]
        if missing:
            msg = json.dumps(
                {"ok": False, "error": f"존재하지 않는 폴더: {', '.join(missing)}"},
                ensure_ascii=False).encode("utf-8")
            self._serve_bytes(msg, "application/json; charset=utf-8")
            return

        # M-1(P107): heal_count 리셋은 RESETTABLE 상태에서만 (_post_run_merge와 동일 정책).
        # quick.json status=heal_needed(힐링 재실행) 시 리셋하면 MAX_HEAL 가드 무력화 → 무한루프.
        # M-4(P121): RESETTABLE_PARALLEL_STATUSES 단일 소스 사용.
        if QUICK_STATE_PATH.exists():
            _quick_status = (load_json(QUICK_STATE_PATH) or {}).get("status", "")
            if _quick_status in RESETTABLE_PARALLEL_STATUSES:
                _safe_update_json(QUICK_STATE_PATH, lambda s: {**s, "heal_count": 0})

        log_path = LOGS_DIR / "quick_run.txt"
        merge_script = PROJECT_ROOT / "parallel" / "99_merge.py"
        log_file = open(log_path, "w", encoding="utf-8")
        no_heal = body.get("no_heal", False)
        cmd = [PYTHON_EXE, "-u", str(merge_script), "--quick", "--group"] + groups
        if no_heal:
            cmd.append("--no-heal")
        proc = sp.Popen(
            cmd, cwd=str(PROJECT_ROOT),
            stdout=log_file, stderr=sp.STDOUT,
        )
        _register_spawned_proc(proc, tag="run_quick")  # P77: tag 등록
        log_file.close()
        print(f"[Dashboard] 빠른 실행 (PID: {proc.pid}, groups: {groups})")
        self._serve_bytes(
            json.dumps({"ok": True, "pid": proc.pid,
                         "log": str(log_path)}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8"
        )

    def _post_import_convert(self):
        import openpyxl
        body = _read_body(self)
        fname = body.get("file", "").strip()
        sheet_names = body.get("sheets", [])
        if not fname or not sheet_names:
            self._serve_bytes(
                b'{"ok":false,"error":"file and sheets required"}',
                "application/json; charset=utf-8")
            return
        if not is_safe_filename(fname):
            self._serve_bytes(
                b'{"ok":false,"error":"invalid file parameter"}',
                "application/json; charset=utf-8")
            return
        fpath = IMPORT_DIR / fname
        if not fpath.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"{fname} not found"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
            return
        try:
            wb = openpyxl.load_workbook(str(fpath), data_only=True)
            results = []
            for sn in sheet_names:
                if sn not in wb.sheetnames:
                    results.append({"sheet": sn, "count": 0, "error": "시트 없음"})
                    continue
                cases = _parse_excel_sheet(wb, sn)
                folder_name = re.sub(r'\s+', '_', sn.strip().lower())
                out_dir = TESTCASES_DIR / folder_name
                # 기존 파일 정리
                if out_dir.exists():
                    for old in out_dir.glob("tc_*.md"):
                        old.unlink()
                count = _write_tc_files(cases, out_dir)
                results.append({"sheet": sn, "count": count,
                                "folder": f"testcases/{folder_name}/"})
            wb.close()
            self._serve_bytes(
                json.dumps({"ok": True, "results": results},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")
        except Exception as e:
            self._serve_bytes(
                json.dumps({"ok": False, "error": str(e)},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8")

    def _post_import_preview(self):
        """POST /api/import/preview — dry-run: 실제 파일 수정 없이 변환 결과 미리보기."""
        import hashlib as _hashlib
        import tempfile

        body = _read_body(self)
        file_id   = body.get("file_id", "").strip()
        sheet_name = body.get("sheet_name", "").strip()
        mappings  = body.get("mappings", {})

        # 필수 매핑 검사
        REQUIRED_MAPPINGS = ["tc_id", "title", "steps", "expected"]
        for f in REQUIRED_MAPPINGS:
            if f not in mappings:
                self._serve_bytes(
                    json.dumps({"ok": False, "error": f"필수 매핑 누락: {f}"},
                               ensure_ascii=False).encode("utf-8"),
                    "application/json; charset=utf-8",
                    status=400,
                )
                return

        if not file_id or not sheet_name:
            self._serve_bytes(
                b'{"ok":false,"error":"file_id and sheet_name required"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return

        # file_id로 파일 탐색 (SHA256(filename)[:8])
        file_path: Path | None = None
        if IMPORT_DIR.exists():
            for f in IMPORT_DIR.glob("*.xlsx"):
                if _hashlib.sha256(f.name.encode("utf-8")).hexdigest()[:8] == file_id:
                    file_path = f
                    break

        if file_path is None:
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"file_id '{file_id}' 에 해당하는 파일 없음"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8",
                status=404,
            )
            return

        # Excel 파싱
        try:
            from _excel_import import parse_sheet  # type: ignore[import]
            rows = parse_sheet(file_path, sheet_name, mappings)
        except Exception as exc:
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"Excel 파싱 실패: {exc}"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8",
                status=422,
            )
            return

        # 기존 testcases 로드 및 분류
        from _import_validator import load_existing_testcases, classify_row  # type: ignore[import]
        existing = load_existing_testcases(TESTCASES_DIR)

        result_rows = []
        for row in rows:
            classification = classify_row(row, existing)
            result_rows.append({**row, **classification})

        # summary 집계
        statuses = ["added", "updated", "conflict", "error", "same"]
        summary = {s: sum(1 for r in result_rows if r.get("status") == s) for s in statuses}

        # 세션 ID 생성: sess_{YYYYMMDD}_{SHA256[:6]}
        session_content = json.dumps(result_rows, ensure_ascii=False, sort_keys=True)
        content_hash = _hashlib.sha256(session_content.encode("utf-8")).hexdigest()[:6]
        session_id = f"sess_{datetime.now().strftime('%Y%m%d')}_{content_hash}"

        # 세션 저장 — state/import_sessions/{session_id}.json
        IMPORT_SESSIONS_DIR.mkdir(parents=True, exist_ok=True)
        session_data = {
            "session_id":  session_id,
            "file_id":     file_id,
            "file_name":   file_path.name,
            "sheet_name":  sheet_name,
            "mappings":    mappings,
            "summary":     summary,
            "rows":        result_rows,
            "created_at":  datetime.now().isoformat(),
            "status":      "preview",
        }
        session_path = IMPORT_SESSIONS_DIR / f"{session_id}.json"
        _tmp_fd, _tmp_path = tempfile.mkstemp(dir=IMPORT_SESSIONS_DIR, suffix=".tmp")
        try:
            import os as _os
            with _os.fdopen(_tmp_fd, "w", encoding="utf-8") as _f:
                json.dump(session_data, _f, ensure_ascii=False, indent=2)
            Path(_tmp_path).replace(session_path)
        except Exception:
            Path(_tmp_path).unlink(missing_ok=True)
            raise

        # 응답: _row 같은 내부 필드는 row로 노출
        response_rows = []
        for r in result_rows:
            response_rows.append({
                "row":    r.get("_row", 0),
                "tc_id":  r.get("tc_id", ""),
                "title":  r.get("title", ""),
                "group":  r.get("group", ""),
                "status": r.get("status", ""),
                "reason": r.get("reason", ""),
            })

        self._serve_bytes(
            json.dumps({
                "ok":         True,
                "session_id": session_id,
                "summary":    summary,
                "rows":       response_rows,
            }, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    # ── Import Studio run-based API (S5-S8) ──────────────────────
    def _import_error_v2(self, exc, default_status: int = 400):
        code = getattr(exc, "code", "IMPORT_ERROR")
        status = {
            "FILE_NOT_FOUND": 404,
            "RUN_NOT_FOUND": 404,
            "SNAPSHOT_NOT_FOUND": 404,
            "ALREADY_COMMITTED": 409,
            "IDEMPOTENCY_CONFLICT": 409,
            "COMMIT_LOCKED": 409,
            "SOURCE_CHANGED": 409,
            "TARGET_CHANGED": 409,
            "UNRESOLVED_CONFLICT": 409,
            "ROLLBACK_CONFLICT": 409,
            "RECOVERY_CONFLICT": 409,
            "PROFILE_NOT_FOUND": 404,
            "PROFILE_EXISTS": 409,
            "PROFILE_STORE_ERROR": 500,
            "RUN_CORRUPT": 500,
            "SNAPSHOT_CORRUPT": 500,
            "COMMIT_RECOVERED": 500,
            "ROLLBACK_VERIFICATION_FAILED": 500,
        }.get(code, default_status)
        self._serve_bytes(
            json.dumps({"ok": False, "error": str(exc), "code": code},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8", status=status,
        )

    def _post_import_preview_v2(self):
        from _import_commit import ImportRunError, create_preview  # type: ignore[import]

        try:
            run = create_preview(
                _read_body(self), IMPORT_DIR, TESTCASES_DIR, IMPORT_SESSIONS_DIR,
            )
        except ImportRunError as exc:
            self._import_error_v2(exc)
            return
        except Exception as exc:
            self._import_error_v2(
                ImportRunError(f"Excel 파싱 실패: {exc}", "PARSE_FAILED"), 422
            )
            return

        rows = [{
            "source_file_id": row.get("_source_file_id", ""),
            "file_name": row.get("_source_file", ""),
            "sheet_name": row.get("_source_sheet", ""),
            "source_row": row.get("_row", 0),
            "row": row.get("_row", 0),
            "tc_id": row.get("tc_id", ""),
            "title": row.get("title", ""),
            "group": row.get("group", ""),
            "status": row.get("status", ""),
            "reason_code": row.get("reason_code", ""),
            "reason": row.get("reason", ""),
            "excluded": row.get("excluded", False),
            "decision": row.get("decision", "automatic"),
            "before": row.get("before"),
            "after": row.get("after"),
        } for row in run["rows"]]
        self._serve_bytes(
            json.dumps({
                "ok": True,
                "run_id": run["run_id"],
                "session_id": run["run_id"],
                "status": run["status"],
                "sources": run["sources"],
                "summary": run["summary"],
                "rows": rows,
            }, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_commit_v2(self):
        from _import_commit import ImportRunError, commit_run  # type: ignore[import]

        body = _read_body(self)
        run_id = str(body.get("run_id") or body.get("session_id") or "").strip()
        idempotency_key = str(body.get("idempotency_key") or "").strip()
        decisions = body.get("decisions", [])
        policy = str(body.get("policy") or "skip-conflict").strip()
        if policy not in {"skip-conflict", "overwrite", "replace-with-snapshot"}:
            policy = "skip-conflict"
        if not isinstance(decisions, list):
            self._import_error_v2(ImportRunError("decisions must be an array", "INVALID_DECISIONS"))
            return
        if len(idempotency_key) > 128:
            self._import_error_v2(ImportRunError("idempotency_key too long", "INVALID_REQUEST"))
            return
        if not run_id:
            self._import_error_v2(ImportRunError("run_id required", "INVALID_REQUEST"))
            return
        try:
            result = commit_run(
                run_id, IMPORT_DIR, TESTCASES_DIR, IMPORT_SESSIONS_DIR,
                IMPORT_SNAPSHOTS_DIR, PROJECT_ROOT, idempotency_key, decisions, policy,
            )
        except ImportRunError as exc:
            self._import_error_v2(exc)
            return
        self._serve_bytes(
            json.dumps({"ok": True, **result}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_rollback_v2(self):
        from _import_commit import ImportRunError, rollback_run  # type: ignore[import]

        body = _read_body(self)
        run_id = str(body.get("run_id") or "").strip()
        # Legacy clients know only snapshot_id; resolve it to its owning run.
        if not run_id and body.get("snapshot_id"):
            snap_id = str(body["snapshot_id"])
            if not is_safe_filename(snap_id):
                self._import_error_v2(ImportRunError("invalid snapshot_id", "INVALID_REQUEST"))
                return
            snaps_root = IMPORT_SNAPSHOTS_DIR.resolve()
            manifest = (snaps_root / snap_id / "manifest.json").resolve()
            if not manifest.is_relative_to(snaps_root):
                self._import_error_v2(ImportRunError("invalid snapshot_id", "INVALID_REQUEST"))
                return
            if manifest.exists():
                run_id = str((load_json(manifest) or {}).get("run_id", ""))
        if not run_id:
            self._import_error_v2(ImportRunError("run_id required", "INVALID_REQUEST"))
            return
        try:
            result = rollback_run(
                run_id, IMPORT_SESSIONS_DIR, IMPORT_SNAPSHOTS_DIR, PROJECT_ROOT,
                TESTCASES_DIR,
            )
        except ImportRunError as exc:
            self._import_error_v2(exc)
            return
        self._serve_bytes(
            json.dumps({"ok": True, **result}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_run_rollback_v2(self, path: str):
        from _import_commit import ImportRunError, rollback_run  # type: ignore[import]

        run_id = path.removeprefix("/api/import/runs/").removesuffix("/rollback").strip("/")
        try:
            result = rollback_run(
                run_id, IMPORT_SESSIONS_DIR, IMPORT_SNAPSHOTS_DIR, PROJECT_ROOT,
                TESTCASES_DIR,
            )
        except ImportRunError as exc:
            self._import_error_v2(exc)
            return
        self._serve_bytes(
            json.dumps({"ok": True, **result}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _get_import_run_v2(self, path: str):
        from _import_commit import ImportRunError, load_run, rows_csv  # type: ignore[import]

        relative = path.removeprefix("/api/import/runs/").strip("/")
        skipped_csv = relative.endswith("/skipped.csv")
        run_id = relative.removesuffix("/skipped.csv") if skipped_csv else relative
        try:
            run = load_run(IMPORT_SESSIONS_DIR, run_id)
        except ImportRunError as exc:
            self._import_error_v2(exc)
            return
        if skipped_csv:
            payload = rows_csv(run, skipped_only=True)
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", f'attachment; filename="import_skipped_{run_id}.csv"')
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return
        # The durable representation is also the result contract. Internal
        # underscore keys retain provenance without leaking filesystem paths.
        public = {**run, "rows": [{
            **{k: v for k, v in row.items() if not k.startswith("_")},
            "file_name": row.get("_source_file", ""),
            "source_file_id": row.get("_source_file_id", ""),
            "sheet_name": row.get("_source_sheet", ""),
            "source_row": row.get("_row", 0),
        } for row in run.get("rows", [])]}
        self._serve_bytes(
            json.dumps({"ok": True, **public}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _delete_import_profile_v2(self, profile_id: str):
        from _paths import IMPORT_PROFILES_PATH
        from _import_commit import ImportRunError

        if not is_safe_filename(profile_id):
            self._serve_bytes(b'{"ok":false,"error":"invalid profile id","code":"INVALID_PROFILE_ID"}',
                              "application/json; charset=utf-8", status=400)
            return
        try:
            def mutate(data):
                profiles = data.get("profiles", [])
                remaining = [p for p in profiles if str(p.get("id")) != profile_id]
                if len(remaining) == len(profiles):
                    raise ImportRunError("profile not found", "PROFILE_NOT_FOUND")
                data["profiles"] = remaining
                return profile_id
            _update_profiles_locked(IMPORT_PROFILES_PATH, mutate)
        except ImportRunError as exc:
            self._import_error_v2(exc, 404)
            return
        self._serve_bytes(
            json.dumps({"ok": True, "deleted": profile_id}).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _put_import_profile_v2(self, profile_id: str):
        from _paths import IMPORT_PROFILES_PATH
        from _import_commit import ImportRunError

        if not is_safe_filename(profile_id):
            self._serve_bytes(b'{"ok":false,"error":"invalid profile id","code":"INVALID_PROFILE_ID"}',
                              "application/json; charset=utf-8", status=400)
            return
        body = _read_body(self)
        if not set(body).intersection({"name", "mappings"}):
            self._serve_bytes(b'{"ok":false,"error":"name or mappings required","code":"INVALID_PROFILE"}',
                              "application/json; charset=utf-8", status=400)
            return
        if "name" in body:
            name = str(body["name"]).strip()
            if not name or len(name) > 50:
                self._serve_bytes(b'{"ok":false,"error":"invalid name","code":"INVALID_PROFILE"}',
                                  "application/json; charset=utf-8", status=400)
                return
        if "mappings" in body:
            if (not isinstance(body["mappings"], dict) or not body["mappings"]
                    or any(not isinstance(key, str) or not isinstance(value, str)
                           for key, value in body["mappings"].items())):
                self._serve_bytes(b'{"ok":false,"error":"invalid mappings","code":"INVALID_PROFILE"}',
                                  "application/json; charset=utf-8", status=400)
                return
        try:
            def mutate(data):
                profile = next((p for p in data.get("profiles", []) if str(p.get("id")) == profile_id), None)
                if profile is None:
                    raise ImportRunError("profile not found", "PROFILE_NOT_FOUND")
                if "name" in body:
                    name = str(body["name"]).strip()
                    if any(p is not profile and p.get("name") == name for p in data["profiles"]):
                        raise ImportRunError("profile name already exists", "PROFILE_EXISTS")
                    profile["name"] = name
                if "mappings" in body:
                    profile["mappings"] = body["mappings"]
                profile["updated_at"] = datetime.now().isoformat()
                return profile
            profile = _update_profiles_locked(IMPORT_PROFILES_PATH, mutate)
        except ImportRunError as exc:
            self._import_error_v2(exc, 404 if exc.code == "PROFILE_NOT_FOUND" else 409)
            return
        self._serve_bytes(
            json.dumps({"ok": True, "profile": profile}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _get_import_profiles_v2(self):
        from _import_commit import ImportRunError
        from _paths import IMPORT_PROFILES_PATH
        try:
            data = _read_profiles_locked(IMPORT_PROFILES_PATH)
        except ImportRunError as exc:
            self._import_error_v2(exc, 500)
            return
        self._serve_bytes(
            json.dumps({"ok": True, "profiles": data.get("profiles", [])},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_profiles_v2(self):
        import secrets
        from _paths import IMPORT_PROFILES_PATH
        from _import_commit import ImportRunError

        body = _read_body(self)
        name = str(body.get("name", "")).strip()
        mappings = body.get("mappings")
        if (not name or len(name) > 50 or not isinstance(mappings, dict) or not mappings
                or any(not isinstance(key, str) or not isinstance(value, str)
                       for key, value in mappings.items())):
            self._serve_bytes(b'{"ok":false,"error":"invalid name or mappings","code":"INVALID_PROFILE"}',
                              "application/json; charset=utf-8", status=400)
            return
        try:
            def mutate(data):
                profiles = data.setdefault("profiles", [])
                if any(profile.get("name") == name for profile in profiles):
                    raise ImportRunError("profile name already exists", "PROFILE_EXISTS")
                profile = {"id": f"prof_{secrets.token_hex(4)}", "name": name,
                           "mappings": mappings, "created_at": datetime.now().isoformat()}
                profiles.append(profile)
                return profile
            profile = _update_profiles_locked(IMPORT_PROFILES_PATH, mutate)
        except ImportRunError as exc:
            self._import_error_v2(exc, 409)
            return
        self._serve_bytes(
            json.dumps({"ok": True, "id": profile["id"], "profile": profile},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8", status=201,
        )

    def _get_import_profiles(self):
        """GET /api/import/profiles — 매핑 프로필 목록."""
        from _paths import IMPORT_PROFILES_PATH
        data = load_json(IMPORT_PROFILES_PATH) or {"profiles": []}
        self._serve_bytes(
            json.dumps({"ok": True, "profiles": data.get("profiles", [])},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_profiles(self):
        """POST /api/import/profiles — 매핑 프로필 저장."""
        import secrets
        import tempfile
        from _paths import IMPORT_PROFILES_PATH

        body = _read_body(self)
        name     = body.get("name", "").strip()
        mappings = body.get("mappings", {})
        if not name or not mappings:
            self._serve_bytes(
                b'{"ok":false,"error":"name and mappings required"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return
        if len(name) > 50:
            self._serve_bytes(
                b'{"ok":false,"error":"name must be 50 characters or fewer"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return
        if not isinstance(mappings, dict) or any(not isinstance(v, str) for v in mappings.values()):
            self._serve_bytes(
                b'{"ok":false,"error":"mappings values must be strings"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return

        IMPORT_PROFILES_PATH.parent.mkdir(parents=True, exist_ok=True)
        data = load_json(IMPORT_PROFILES_PATH) or {"profiles": []}
        profiles: list = data.get("profiles", [])

        # 동일 이름 중복 검사
        if any(p.get("name") == name for p in profiles):
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"프로필 '{name}' 이미 존재"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8",
                status=409,
            )
            return

        profile_id = f"prof_{secrets.token_hex(4)}"
        new_profile = {
            "id":         profile_id,
            "name":       name,
            "mappings":   mappings,
            "created_at": datetime.now().isoformat(),
        }
        profiles.append(new_profile)
        data["profiles"] = profiles

        _tmp_fd, _tmp_path = tempfile.mkstemp(dir=IMPORT_PROFILES_PATH.parent, suffix=".tmp")
        try:
            import os as _os
            with _os.fdopen(_tmp_fd, "w", encoding="utf-8") as _f:
                json.dump(data, _f, ensure_ascii=False, indent=2)
            Path(_tmp_path).replace(IMPORT_PROFILES_PATH)
        except Exception:
            Path(_tmp_path).unlink(missing_ok=True)
            raise

        self._serve_bytes(
            json.dumps({"ok": True, "id": profile_id}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
            status=201,
        )

    def _post_import_profiles_update(self):
        """POST /api/import/profiles/update — 프로필 이름 또는 매핑 수정."""
        from _paths import IMPORT_PROFILES_PATH
        from _import_commit import ImportRunError

        body = _read_body(self)
        profile_id = body.get("id", "").strip()
        new_name   = body.get("name", "").strip()
        new_mappings = body.get("mappings")

        if not profile_id:
            self._serve_bytes(
                b'{"ok":false,"error":"id required"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return

        # name 검증
        if new_name and len(new_name) > 50:
            self._serve_bytes(
                b'{"ok":false,"error":"name must be 50 characters or fewer"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return

        # mappings 타입 검증
        if new_mappings is not None:
            if not isinstance(new_mappings, dict) or any(
                not isinstance(v, str) for v in new_mappings.values()
            ):
                self._serve_bytes(
                    b'{"ok":false,"error":"mappings values must be strings"}',
                    "application/json; charset=utf-8",
                    status=400,
                )
                return

        try:
            def mutate(data):
                profiles = data.setdefault("profiles", [])
                target = next((p for p in profiles if p.get("id") == profile_id), None)
                if target is None:
                    raise ImportRunError(f"프로필 '{profile_id}' 없음", "PROFILE_NOT_FOUND")
                if new_name and any(
                    p.get("name") == new_name and p.get("id") != profile_id
                    for p in profiles
                ):
                    raise ImportRunError(f"프로필 이름 '{new_name}' 이미 존재", "PROFILE_EXISTS")
                if new_name:
                    target["name"] = new_name
                if new_mappings is not None:
                    target["mappings"] = new_mappings
                target["updated_at"] = datetime.now().isoformat()
                return dict(target)

            updated = _update_profiles_locked(IMPORT_PROFILES_PATH, mutate)
        except ImportRunError as exc:
            self._import_error_v2(exc, 404 if exc.code == "PROFILE_NOT_FOUND" else 409)
            return

        self._serve_bytes(
            json.dumps({"ok": True, "id": profile_id, "profile": updated},
                       ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_profiles_delete(self):
        """POST /api/import/profiles/delete — 프로필 삭제 (DELETE 대안)."""
        from _paths import IMPORT_PROFILES_PATH
        from _import_commit import ImportRunError

        body = _read_body(self)
        profile_id = body.get("id", "").strip()
        if not profile_id:
            self._serve_bytes(
                b'{"ok":false,"error":"id required"}',
                "application/json; charset=utf-8",
                status=400,
            )
            return

        try:
            def mutate(data):
                profiles = data.setdefault("profiles", [])
                remaining = [p for p in profiles if p.get("id") != profile_id]
                if len(remaining) == len(profiles):
                    raise ImportRunError(f"프로필 '{profile_id}' 없음", "PROFILE_NOT_FOUND")
                data["profiles"] = remaining
                return profile_id

            _update_profiles_locked(IMPORT_PROFILES_PATH, mutate)
        except ImportRunError as exc:
            self._import_error_v2(exc, 404)
            return

        self._serve_bytes(
            json.dumps({"ok": True, "deleted": profile_id}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_commit(self):
        """POST /api/import/commit — preview 세션 결과를 실제 tc_*.md 파일에 반영."""
        import shutil
        import tempfile
        import os as _os
        from _import_validator import load_existing_testcases  # type: ignore[import]

        body = _read_body(self)
        session_id = body.get("session_id", "").strip()

        if not session_id:
            self._serve_bytes(
                b'{"ok":false,"error":"session_id required"}',
                "application/json; charset=utf-8", status=400)
            return

        if not is_safe_filename(session_id):
            self._serve_bytes(
                json.dumps({"ok": False, "error": "invalid session_id"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=400)
            return

        # 경로 봉쇄 (P66 패턴)
        _sessions_root = IMPORT_SESSIONS_DIR.resolve()
        session_path = (_sessions_root / f"{session_id}.json").resolve()
        if not session_path.is_relative_to(_sessions_root):
            self._serve_bytes(
                b'{"ok":false,"error":"invalid session_id"}',
                "application/json; charset=utf-8", status=400)
            return

        if not session_path.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"세션 '{session_id}' 없음 또는 만료"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=400)
            return

        session_data = load_json(session_path) or {}

        # 이미 커밋된 세션 → 409
        if session_data.get("status") == "committed":
            self._serve_bytes(
                json.dumps({"ok": False, "error": "이미 커밋된 세션",
                            "code": "ALREADY_COMMITTED"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=409)
            return

        rows = session_data.get("rows", [])
        commit_rows = [r for r in rows if r.get("status") in ("added", "updated")]
        skipped = len(rows) - len(commit_rows)

        # 스냅샷 ID: snap_{YYYYMMDD}_{HHMMSS}
        snap_id = f"snap_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        IMPORT_SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)
        _snaps_root = IMPORT_SNAPSHOTS_DIR.resolve()
        snap_dir = (_snaps_root / snap_id).resolve()
        if not snap_dir.is_relative_to(_snaps_root):
            self._serve_bytes(b'{"ok":false,"error":"snapshot path error"}',
                             "application/json; charset=utf-8", status=500)
            return
        snap_dir.mkdir(parents=True, exist_ok=True)

        # 기존 tc_*.md 백업 (updated 케이스)
        _tc_root = TESTCASES_DIR.resolve()
        _proj_root = PROJECT_ROOT.resolve()
        existing = load_existing_testcases(TESTCASES_DIR)
        for r in commit_rows:
            tc_id = r.get("tc_id", "").strip()
            if r.get("status") == "updated" and tc_id in existing:
                orig_path = existing[tc_id]["path"]  # type: ignore[index]
                try:
                    rel = orig_path.resolve().relative_to(_proj_root)
                    backup_path = (snap_dir / rel).resolve()
                    if backup_path.is_relative_to(snap_dir):
                        backup_path.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(orig_path, backup_path)
                except Exception:
                    pass

        # tc_*.md 파일 생성/갱신
        committed = 0
        for r in commit_rows:
            tc_id   = r.get("tc_id", "").strip()
            group   = r.get("group", "").strip()
            title   = r.get("title", "").strip()
            steps   = r.get("steps", "").strip()
            expected = r.get("expected", "").strip()
            priority = (r.get("priority", "") or "medium").strip()
            tags    = r.get("tags", [])

            if not tc_id or not group:
                continue
            if not is_valid_group_name(group):
                continue

            group_dir = (_tc_root / group).resolve()
            if not group_dir.is_relative_to(_tc_root):
                continue
            group_dir.mkdir(parents=True, exist_ok=True)

            # 파일명: tc_{tc_id}_{slug}.md — tc_*.md glob 패턴 준수
            slug = re.sub(r'\W+', '_', title.lower())[:40].strip('_') or "case"
            tc_filename = f"tc_{tc_id}_{slug}.md"
            md_path = (group_dir / tc_filename).resolve()
            if not md_path.is_relative_to(group_dir):
                continue

            tags_str = ", ".join(str(t) for t in tags) if tags else "general"
            content = (
                f"---\nid: {tc_id}\ndata_key: null\npriority: {priority}\n"
                f"tags: [{tags_str}]\ntype: structured\n---\n"
                f"# {title}\n\n"
                f"## Steps\n{steps}\n\n"
                f"## Expected\n{expected}\n"
            )

            _tmp_fd, _tmp_path = tempfile.mkstemp(dir=group_dir, suffix=".tmp")
            try:
                with _os.fdopen(_tmp_fd, "w", encoding="utf-8") as _f:
                    _f.write(content)
                Path(_tmp_path).replace(md_path)
                committed += 1
            except Exception:
                Path(_tmp_path).unlink(missing_ok=True)

        # 세션 상태를 "committed"로 마킹
        session_data["status"] = "committed"
        session_data["snapshot_id"] = snap_id
        session_data["committed_at"] = datetime.now().isoformat()
        _s_tmp_fd, _s_tmp_path = tempfile.mkstemp(dir=IMPORT_SESSIONS_DIR, suffix=".tmp")
        try:
            with _os.fdopen(_s_tmp_fd, "w", encoding="utf-8") as _f:
                json.dump(session_data, _f, ensure_ascii=False, indent=2)
            Path(_s_tmp_path).replace(session_path)
        except Exception:
            Path(_s_tmp_path).unlink(missing_ok=True)

        self._serve_bytes(
            json.dumps({
                "ok":          True,
                "snapshot_id": snap_id,
                "committed":   committed,
                "skipped":     skipped,
            }, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _post_import_rollback(self):
        """POST /api/import/rollback — 스냅샷으로 파일 복원."""
        import shutil

        body = _read_body(self)
        snap_id = body.get("snapshot_id", "").strip()

        if not snap_id:
            self._serve_bytes(
                b'{"ok":false,"error":"snapshot_id required"}',
                "application/json; charset=utf-8", status=400)
            return

        # is_safe_filename() 검증 필수
        if not is_safe_filename(snap_id):
            self._serve_bytes(
                json.dumps({"ok": False, "error": "invalid snapshot_id"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=400)
            return

        # 경로 봉쇄 (P66 패턴)
        _snaps_root = IMPORT_SNAPSHOTS_DIR.resolve()
        snap_dir = (_snaps_root / snap_id).resolve()
        if not snap_dir.is_relative_to(_snaps_root):
            self._serve_bytes(
                b'{"ok":false,"error":"invalid snapshot_id"}',
                "application/json; charset=utf-8", status=400)
            return

        if not snap_dir.exists() or not snap_dir.is_dir():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"스냅샷 '{snap_id}' 없음"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=404)
            return

        # 스냅샷 내 파일을 원래 경로로 복원
        _proj_root = PROJECT_ROOT.resolve()
        restored = 0
        for f in snap_dir.rglob("*.md"):
            f_resolved = f.resolve()
            snap_dir_resolved = snap_dir.resolve()
            if not f_resolved.is_relative_to(snap_dir_resolved):
                continue
            rel = f_resolved.relative_to(snap_dir_resolved)
            orig_path = (_proj_root / rel).resolve()
            # 경로 탈출 방지
            if not orig_path.is_relative_to(_proj_root):
                continue
            orig_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(f, orig_path)
            restored += 1

        self._serve_bytes(
            json.dumps({"ok": True, "restored": restored}, ensure_ascii=False).encode("utf-8"),
            "application/json; charset=utf-8",
        )

    def _get_import_preview_csv(self):
        """GET /api/import/preview/csv?session_id=... — CSV 다운로드."""
        import csv
        import io
        from _import_commit import sanitize_csv_cell
        from urllib.parse import urlparse, parse_qs

        qs = parse_qs(urlparse(self.path).query)
        session_id = qs.get("session_id", [""])[0]

        if not session_id:
            self._serve_bytes(
                b'{"ok":false,"error":"session_id parameter required"}',
                "application/json; charset=utf-8", status=400)
            return

        if not is_safe_filename(session_id):
            self._serve_bytes(
                json.dumps({"ok": False, "error": "invalid session_id"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=400)
            return

        # 경로 봉쇄 (P66 패턴)
        _sessions_root = IMPORT_SESSIONS_DIR.resolve()
        session_path = (_sessions_root / f"{session_id}.json").resolve()
        if not session_path.is_relative_to(_sessions_root):
            self._serve_bytes(
                b'{"ok":false,"error":"invalid session_id"}',
                "application/json; charset=utf-8", status=400)
            return

        if not session_path.exists():
            self._serve_bytes(
                json.dumps({"ok": False, "error": f"세션 '{session_id}' 없음 또는 만료"},
                           ensure_ascii=False).encode("utf-8"),
                "application/json; charset=utf-8", status=404)
            return

        session_data = load_json(session_path) or {}
        rows = session_data.get("rows", [])

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["row", "tc_id", "title", "group", "status", "reason"])
        for r in rows:
            writer.writerow([sanitize_csv_cell(value) for value in [
                r.get("_row", r.get("row", "")),
                r.get("tc_id", ""),
                r.get("title", ""),
                r.get("group", ""),
                r.get("status", ""),
                r.get("reason", ""),
            ]])

        csv_bytes = buf.getvalue().encode("utf-8")
        fname = f"import_preview_{session_id}.csv"
        self.send_response(200)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
        self.send_header("Access-Control-Allow-Origin", ALLOWED_ORIGIN)
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Content-Length", str(len(csv_bytes)))
        self.end_headers()
        self.wfile.write(csv_bytes)

    # ── Infrastructure helpers ────────────────────────────────────
    def _serve_sse(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/event-stream; charset=utf-8")
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Access-Control-Allow-Origin", ALLOWED_ORIGIN)
        self.send_header("X-Accel-Buffering", "no")
        self.end_headers()

        q: queue.Queue = queue.Queue(maxsize=10)
        with _sse_lock:
            _sse_clients.append(q)

        def send(data: str):
            self.wfile.write(f"data: {data}\n\n".encode("utf-8"))
            self.wfile.flush()

        try:
            # 연결 즉시 현재 상태 전송
            send(json.dumps(build_dialogs(), ensure_ascii=False))
            while True:
                try:
                    q.get(timeout=15)
                    send(json.dumps(build_dialogs(), ensure_ascii=False))
                except queue.Empty:
                    # keepalive
                    self.wfile.write(b": ping\n\n")
                    self.wfile.flush()
        except Exception:
            pass
        finally:
            with _sse_lock:
                if q in _sse_clients:
                    _sse_clients.remove(q)

    def _serve_file(self, path: Path, content_type: str):
        if path.exists():
            content = path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(content)))
            if "html" in content_type:
                self.send_header("X-XSS-Protection", "0")
            self.end_headers()
            self.wfile.write(content)
        else:
            self.send_response(404)
            self.end_headers()

    def _serve_json(self, path: Path):
        content = path.read_bytes() if path.exists() else b'{"pipeline_url":"","started_at":"","sessions":[]}'
        self._serve_bytes(content, "application/json; charset=utf-8")

    def _serve_bytes(self, content: bytes, content_type: str, status: int = 200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Access-Control-Allow-Origin", ALLOWED_ORIGIN)
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {format % args}")


class ReusableHTTPServer(ThreadingHTTPServer):
    allow_reuse_address = True
    allow_reuse_port = True


def _is_port_in_use(port: int, host: str = "127.0.0.1") -> bool:
    import socket
    check_host = "127.0.0.1" if host in ("", "0.0.0.0") else host
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        return s.connect_ex((check_host, port)) == 0


def main():
    parser = argparse.ArgumentParser(description="QA Agent Dashboard Server")
    parser.add_argument(
        "--host", default="127.0.0.1", help="bind host (default: 127.0.0.1)"
    )
    parser.add_argument(
        "--port", type=int, default=PORT, help=f"port (default: {PORT})"
    )
    args = parser.parse_args()
    host = args.host
    port = args.port

    display_host = (
        "localhost" if host in ("", "0.0.0.0", "127.0.0.1") else host
    )
    global ALLOWED_HOSTS, ALLOWED_ORIGIN
    if not os.environ.get("ALLOWED_HOSTS") and (host != "127.0.0.1" or port != PORT):
        ALLOWED_HOSTS = {
            f"localhost:{port}",
            f"127.0.0.1:{port}",
            f"{host}:{port}",
        }
    if not os.environ.get("ALLOWED_ORIGIN") and (host != "127.0.0.1" or port != PORT):
        ALLOWED_ORIGIN = f"http://{display_host}:{port}"

    if _is_port_in_use(port, host):
        url = f"http://{display_host}:{port}"
        print(f"[Dashboard] 이미 실행 중: {url}")
        webbrowser.open(url)
        return
    server = ReusableHTTPServer((host, port), DashboardHandler)
    url = f"http://{display_host}:{port}"
    print(f"[Dashboard] 서버 시작: {url}")
    print("[Dashboard] 종료: Ctrl+C")
    webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Dashboard] 서버 종료")


if __name__ == "__main__":
    main()
